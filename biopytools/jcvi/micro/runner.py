"""
微观共线性可视化运行器|Microsynteny Visualization Runner

流程:
1. 检测/自动运行mcscan管道 (JcviPipeline)
2. jcvi.compara.synteny mcscan → .blocks
3. 基于BED坐标筛选区域内的block
4. 按基因列表过滤block行
5. 生成过滤后的merged.bed
6. 生成layout
7. jcvi.graphics.synteny → PDF
"""

import os
import subprocess
import time
from pathlib import Path
from datetime import datetime
from typing import Optional, List, Tuple, Dict, Set

from .config import MicroConfig
from ..utils import JcviLogger, build_jcvi_command, get_jcvi_stem
from ..pipeline import JcviPipeline


class MicroRunner:
    """微观共线性可视化运行器|Microsynteny Visualization Runner"""

    VERSION = "2.0.0"

    def __init__(self, config: MicroConfig, logger: Optional[JcviLogger] = None):
        self.config = config
        self.logger_obj = logger
        self.logger = logger.get_logger() if logger else None
        self.start_time = None

    def run(self) -> bool:
        try:
            self.start_time = time.time()

            if self.logger_obj is None:
                log_dir = Path(self.config.output_dir) / "99_logs"
                log_dir.mkdir(parents=True, exist_ok=True)
                self.logger_obj = JcviLogger(log_dir / "micro.log", "MicroSynteny")
                self.logger = self.logger_obj.get_logger()

            self._print_header()
            self.config.validate()

            pair_parts = self.config.pairs[0].split(",")
            if len(pair_parts) != 2:
                self.logger.error(
                    f"pairs格式错误, 应为 A,B|Invalid pairs format, expected A,B: "
                    f"{self.config.pairs[0]}"
                )
                return False
            name_a, name_b = pair_parts[0].strip(), pair_parts[1].strip()
            self._current_name_a = name_a
            self._current_name_b = name_b

            micro_dir = Path(self.config.output_dir) / "04_micro"
            micro_dir.mkdir(parents=True, exist_ok=True)

            # 步骤1: 确保mcscan管道已运行
            self.logger.info(
                "  步骤1/5: 检测/运行mcscan管道|Step 1/5: Ensure mcscan pipeline"
            )
            pair_dir = self._ensure_mcscan(name_a, name_b)
            if not pair_dir:
                return False

            stem_a, stem_b = get_jcvi_stem(name_a), get_jcvi_stem(name_b)
            pprefix = f"{stem_a}.{stem_b}"
            anchors_file = self._find_anchors_file(pair_dir, pprefix)
            if not anchors_file:
                self.logger.error(
                    f"未找到anchors文件|anchors file not found in: {pair_dir}"
                )
                return False

            # 查找BED文件
            bed_a = self._find_bed(name_a)
            bed_b = self._find_bed(name_b)
            if not bed_a or not bed_b:
                self.logger.error("未找到BED文件|BED files not found")
                return False

            # 步骤2: 生成blocks文件
            self.logger.info(
                "  步骤2/5: 生成blocks|Step 2/5: Generate blocks"
            )
            blocks_file = self._mcscan_blocks(pair_dir, bed_a, pprefix, micro_dir)
            if not blocks_file:
                return False

            # 步骤3: 解析BED + 筛选区域block + 过滤基因
            self.logger.info(
                "  步骤3/5: 筛选block并过滤基因|Step 3/5: Filter blocks and genes"
            )
            filtered_blocks_file = self._filter_and_select_blocks(
                blocks_file, bed_a, bed_b, micro_dir
            )
            if not filtered_blocks_file:
                self.logger.error("未找到匹配的block|No matching block found")
                return False

            # 步骤4: 生成过滤后的merged.bed
            self.logger.info(
                "  步骤4/5: 生成merged.bed|Step 4/5: Generate merged BED"
            )
            merged_bed = self._create_filtered_merged_bed(
                bed_a, bed_b, filtered_blocks_file, micro_dir
            )
            if not merged_bed:
                return False

            # 步骤4.5: 输出共线性配对Excel
            self.logger.info(
                "  步骤4.5/5: 输出共线性配对|Step 4.5/5: Export collinear pairs"
            )
            self._export_pair_excel(filtered_blocks_file, anchors_file,
                                    bed_a, bed_b, micro_dir)

            # 步骤5: 生成layout并绘图
            self.logger.info(
                "  步骤5/5: 绘制microsynteny|Step 5/5: Plot microsynteny"
            )
            layout_file = self._generate_layout(micro_dir, name_a, name_b)
            pdf_labeled = self._plot_synteny(
                filtered_blocks_file, merged_bed, layout_file, micro_dir,
                suffix="labeled", genelabelsize=6,
            )
            pdf_clean = self._plot_synteny(
                filtered_blocks_file, merged_bed, layout_file, micro_dir,
                suffix="clean", genelabelsize=0,
            )

            if pdf_labeled:
                self.logger.info(f"\n  输出文件(有标签)|Output (labeled): {pdf_labeled}")
            if pdf_clean:
                self.logger.info(f"  输出文件(无标签)|Output (clean): {pdf_clean}")

            self._print_footer(pdf_labeled is not None or pdf_clean is not None)
            return pdf_labeled is not None or pdf_clean is not None

        except Exception as e:
            if self.logger:
                self.logger.error(f"程序执行出错|Error: {e}", exc_info=True)
            return False

    # ---- 步骤实现 ----

    def _find_anchors_file(self, pair_dir: Path, pprefix: str) -> Optional[Path]:
        """查找anchors输出文件(兼容liftover开启/关闭)|Find anchors output file"""
        for suffix in [".lifted.anchors", ".anchors"]:
            f = pair_dir / f"{pprefix}{suffix}"
            if f.exists() and f.stat().st_size > 0:
                return f
        return None

    def _ensure_mcscan(self, name_a: str, name_b: str) -> Optional[Path]:
        """确保mcscan管道已运行, 自动检测或执行|Ensure mcscan pipeline is complete"""
        pair_dir = Path(self.config.output_dir) / "03_pairwise" / f"{name_a}_vs_{name_b}"
        stem_a, stem_b = get_jcvi_stem(name_a), get_jcvi_stem(name_b)
        pprefix = f"{stem_a}.{stem_b}"

        lifted_anchors = self._find_anchors_file(pair_dir, pprefix)

        if lifted_anchors:
            self.logger.info(
                f"    检测到已有mcscan结果, 复用|"
                f"Existing mcscan results found, reusing: {pair_dir}"
            )
            return pair_dir

        self.logger.info(
            f"    未检测到mcscan结果, 自动运行|"
            f"No mcscan results found, running pipeline automatically"
        )
        pipeline = JcviPipeline(self.config, self.logger_obj)
        result = pipeline.run()

        pair_dir_check = result.pair_dirs.get((name_a, name_b))
        if not pair_dir_check:
            self.logger.error(
                f"mcscan管道运行失败|mcscan pipeline failed for {name_a} vs {name_b}"
            )
            return None

        return pair_dir_check

    def _mcscan_blocks(self, pair_dir: Path, bed_a: Path,
                       pprefix: str, micro_dir: Path) -> Optional[Path]:
        """生成blocks文件|Generate blocks file via jcvi.compara.synteny mcscan"""
        anchors_file = self._find_anchors_file(pair_dir, pprefix)
        if not anchors_file:
            self.logger.error("    未找到anchors文件|anchors file not found")
            return None
        blocks_out = micro_dir / f"{pprefix}.blocks"

        if blocks_out.exists() and blocks_out.stat().st_size > 0:
            self.logger.info(f"    跳过已完成|Skipping completed: mcscan blocks")
            return blocks_out

        cmd = build_jcvi_command("jcvi.compara.synteny", [
            "mcscan",
            str(bed_a.resolve()),
            str(anchors_file.resolve()),
            f"--iter={self.config.iter_count}",
            "-o", str(blocks_out.resolve()),
        ], self.config.conda_env)

        self.logger.info(f"    命令|Command: {' '.join(cmd)}")
        if self._run_cmd(cmd, cwd=str(pair_dir)) != 0:
            self.logger.error("    mcscan blocks 失败|mcscan blocks failed")
            return None

        if not blocks_out.exists() or blocks_out.stat().st_size == 0:
            self.logger.error("    blocks文件为空|blocks file is empty")
            return None

        return blocks_out

    def _filter_and_select_blocks(
        self,
        blocks_file: Path,
        bed_a: Path,
        bed_b: Path,
        micro_dir: Path,
    ) -> Optional[Path]:
        """筛选区域block并过滤基因|Select blocks by region and filter by genes"""
        # 加载BED索引
        bed_index_a = self._load_bed_index(bed_a)
        bed_index_b = self._load_bed_index(bed_b)
        if not bed_index_a or not bed_index_b:
            self.logger.error("BED文件为空或无法解析|BED file is empty or cannot parse")
            return None

        # 从GFF构建转录本同源映射, 用于解析用户基因ID与BED基因ID的对应关系
        transcript_map_a = self._build_transcript_map(name_a=self._current_name_a)
        transcript_map_b = self._build_transcript_map(name_a=self._current_name_b)

        # 解析并解析基因列表: 用户传入的基因ID → BED中实际存在的ID
        genes_a_raw = self._parse_gene_list(self.config.genes_a)
        genes_b_raw = self._parse_gene_list(self.config.genes_b)
        genes_a_resolved = self._resolve_gene_ids(genes_a_raw, bed_index_a, transcript_map_a) if genes_a_raw else None
        genes_b_resolved = self._resolve_gene_ids(genes_b_raw, bed_index_b, transcript_map_b) if genes_b_raw else None

        if genes_a_resolved is not None:
            self.logger.info(
                f"    基因列表A解析|Genes A resolved: "
                f"{len(genes_a_raw)} → {len(genes_a_resolved)} (BED中存在|found in BED)"
            )
        if genes_b_resolved is not None:
            self.logger.info(
                f"    基因列表B解析|Genes B resolved: "
                f"{len(genes_b_raw)} → {len(genes_b_resolved)} (BED中存在|found in BED)"
            )

        # 解析区域
        region_a = self._parse_region(self.config.region_a)
        region_b = self._parse_region(self.config.region_b)
        if not region_a or not region_b:
            return None

        # 查找区域内的基因(用完整BED ID)
        genes_in_region_a = self._find_region_genes(bed_index_a, region_a)
        genes_in_region_b = self._find_region_genes(bed_index_b, region_b)

        self.logger.info(
            f"    区域A {self.config.region_a} 内基因数|"
            f"Genes in region A {self.config.region_a}: {len(genes_in_region_a)}"
        )
        self.logger.info(
            f"    区域B {self.config.region_b} 内基因数|"
            f"Genes in region B {self.config.region_b}: {len(genes_in_region_b)}"
        )

        if not genes_in_region_a:
            self.logger.warning(
                f"区域A内未找到基因|No genes found in region A: {self.config.region_a}"
            )
        if not genes_in_region_b:
            self.logger.warning(
                f"区域B内未找到基因|No genes found in region B: {self.config.region_b}"
            )

        # 解析blocks
        blocks = self._parse_blocks(blocks_file)
        if not blocks:
            self.logger.error("blocks文件为空|blocks file is empty")
            return None

        self.logger.info(f"    blocks总数|Total blocks: {len(blocks)}")

        # 筛选与两个区域都重叠的block
        matching_indices = self._select_region_blocks(
            blocks, genes_in_region_a, genes_in_region_b
        )

        if not matching_indices:
            self.logger.error(
                "未找到与两个区域都重叠的block|"
                "No block overlapping with both regions"
            )
            return None

        self.logger.info(
            f"    匹配区域block数|Region-matching blocks: {len(matching_indices)}"
        )

        # 扩展block范围
        extend = self.config.extend_blocks
        start_idx = max(0, min(matching_indices) - extend)
        end_idx = min(len(blocks), max(matching_indices) + extend + 1)

        self.logger.info(
            f"    扩展后block范围 [{start_idx}, {end_idx})|"
            f"Extended block range [{start_idx}, {end_idx})"
        )

        # 收集所有过滤后的基因ID (用于后续BED过滤, 存完整BED ID)
        all_filtered_genes: Set[str] = set()

        out_file = micro_dir / "filtered.blocks"
        with open(out_file, "w") as fout:
            for i in range(start_idx, end_idx):
                block_lines = blocks[i]
                filtered_lines = self._filter_gene_rows(
                    block_lines, genes_a_resolved, genes_b_resolved
                )

                if not filtered_lines:
                    continue

                for line in filtered_lines:
                    fields = line.split("\t")
                    if fields[0] != ".":
                        all_filtered_genes.add(fields[0])
                    if len(fields) > 1 and fields[1] != ".":
                        all_filtered_genes.add(fields[1])

                fout.write("\n".join(filtered_lines) + "\n")
                if i < end_idx - 1:
                    fout.write("###\n")

            # 追加区域内但非共线性的展示基因(基因列表指定但无共线性对)
            standalone_a = (genes_a_resolved & genes_in_region_a) - all_filtered_genes if genes_a_resolved else set()
            standalone_b = (genes_b_resolved & genes_in_region_b) - all_filtered_genes if genes_b_resolved else set()
            if standalone_a or standalone_b:
                fout.write("###\n")
                for gene_id in sorted(standalone_a):
                    fout.write(f"{gene_id}\t.\n")
                for gene_id in sorted(standalone_b):
                    fout.write(f".\t{gene_id}\n")

        # 所有展示基因 = 共线性 + 非共线性
        all_display_genes = set(all_filtered_genes)
        all_display_genes.update(standalone_a)
        all_display_genes.update(standalone_b)

        if not all_display_genes:
            self.logger.error("过滤后无基因|No genes after filtering")
            return None

        self.logger.info(
            f"    过滤后基因数|Filtered genes: {len(all_display_genes)}"
            f" (共线性|collinear: {len(all_filtered_genes)},"
            f" 非共线性|non-collinear: {len(standalone_a) + len(standalone_b)})"
        )

        # 缓存所有展示基因集合和解析后的基因列表(供Excel输出使用)
        self._display_genes = all_display_genes
        self._genes_a_resolved = genes_a_resolved or set()
        self._genes_b_resolved = genes_b_resolved or set()

        if not out_file.exists() or out_file.stat().st_size == 0:
            return None

        return out_file

    def _create_filtered_merged_bed(
        self,
        bed_a: Path,
        bed_b: Path,
        filtered_blocks_file: Path,
        micro_dir: Path,
    ) -> Optional[Path]:
        """生成merged.bed|Create merged BED"""
        merged_out = micro_dir / "filtered.merged.bed"

        filtered_genes = getattr(self, "_display_genes", set())
        if not filtered_genes:
            self.logger.error("无过滤基因集合|No filtered gene set available")
            return None

        try:
            with open(merged_out, "w") as fout:
                for bed_file in [bed_a, bed_b]:
                    with open(bed_file, "r") as fin:
                        for line in fin:
                            line = line.rstrip("\n")
                            if not line or line.startswith("#"):
                                continue
                            fields = line.split("\t")
                            if len(fields) >= 4:
                                gene_id = fields[3]
                                if gene_id in filtered_genes:
                                    fout.write(line + "\n")
        except Exception as e:
            self.logger.error(f"生成merged.bed失败|Failed to create merged.bed: {e}")
            return None

        if merged_out.stat().st_size == 0:
            self.logger.error("merged.bed为空|merged.bed is empty")
            return None

        return merged_out

    def _export_pair_excel(self, filtered_blocks_file: Path,
                           anchors_file: Path,
                           bed_a: Path, bed_b: Path,
                           micro_dir: Path):
        """输出共线性配对Excel|Export collinear pairs to Excel

        Sheet1: 目标基因相关的共线性对 + 无配对目标基因
        Sheet2: 两个基因组全部共线性对(来自anchors文件), 含染色体/位置/链信息
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        genes_a_resolved = getattr(self, "_genes_a_resolved", set())
        genes_b_resolved = getattr(self, "_genes_b_resolved", set())
        name_a = self._current_name_a
        name_b = self._current_name_b

        # 构建C-score索引
        cscore_map = self._build_cscore_index()

        # 解析filtered.blocks中的共线性对
        all_pairs = []
        seen_a_in_blocks = set()
        seen_b_in_blocks = set()
        with open(filtered_blocks_file, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or line == "###":
                    continue
                fields = line.split("\t")
                if len(fields) < 2:
                    continue
                ga, gb = fields[0], fields[1]
                if ga != "." and gb != ".":
                    score = cscore_map.get((ga, gb), cscore_map.get((gb, ga), None))
                    all_pairs.append((ga, gb, score))
                    seen_a_in_blocks.add(ga)
                    seen_b_in_blocks.add(gb)

        unpaired_a = genes_a_resolved - seen_a_in_blocks
        unpaired_b = genes_b_resolved - seen_b_in_blocks

        # 样式
        header_fill = PatternFill(start_color="4477AA", end_color="4477AA", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        unpaired_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        headers = [f"{name_a}", f"{name_b}", "C-score", f"{name_a}目标", f"{name_b}目标", "配对状态"]

        def _write_headers(ws):
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = thin_border

        def _write_row(ws, row, values, fill=None):
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=v)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center")
                if fill:
                    cell.fill = fill

        def _auto_width(ws, max_row):
            for col_idx in range(1, len(headers) + 1):
                max_len = len(str(ws.cell(row=1, column=col_idx).value))
                for r in range(2, max_row):
                    v = ws.cell(row=r, column=col_idx).value
                    if v:
                        max_len = max(max_len, len(str(v)))
                ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

        def _format_score(score):
            if score is None:
                return "liftover"
            return round(score, 2)

        # --- Sheet 1: 目标基因相关 ---
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Target Genes"
        _write_headers(ws1)

        row = 2
        for ga, gb, score in all_pairs:
            is_a = ga in genes_a_resolved
            is_b = gb in genes_b_resolved
            if not is_a and not is_b:
                continue
            is_target_a = "是" if is_a else "否"
            is_target_b = "是" if is_b else "否"
            status = "双目标" if is_a and is_b else "单目标"
            _write_row(ws1, row, [ga, gb, _format_score(score), is_target_a, is_target_b, status])
            row += 1

        for ga in sorted(unpaired_a):
            _write_row(ws1, row, [ga, "", "", "是", "", "无共线性配对"], fill=unpaired_fill)
            row += 1
        for gb in sorted(unpaired_b):
            _write_row(ws1, row, ["", gb, "", "", "是", "无共线性配对"], fill=unpaired_fill)
            row += 1

        _auto_width(ws1, row)

        # --- 构建BED位置索引: gene_id -> (chr, start, end, strand) ---
        def _parse_bed_index(bed_path: Path) -> dict:
            idx = {}
            with open(bed_path, "r") as f:
                for line in f:
                    fields = line.strip().split("\t")
                    if len(fields) >= 6:
                        idx[fields[3]] = (fields[0], int(fields[1]), int(fields[2]), fields[5])
            return idx

        bed_a_idx = _parse_bed_index(bed_a)
        bed_b_idx = _parse_bed_index(bed_b)

        # --- Sheet 2: 两个基因组全部共线性对(来自anchors文件) ---
        ws2 = wb.create_sheet("All Pairs")
        all_headers = [
            f"{name_a}_gene", f"{name_a}_chr", f"{name_a}_start", f"{name_a}_end", f"{name_a}_strand",
            f"{name_b}_gene", f"{name_b}_chr", f"{name_b}_start", f"{name_b}_end", f"{name_b}_strand",
            "C-score",
        ]
        for col, h in enumerate(all_headers, 1):
            cell = ws2.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        genome_pairs = []
        with open(anchors_file, "r") as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith("#") or line == "###":
                    continue
                fields = line.split("\t")
                if len(fields) < 2:
                    continue
                ga, gb = fields[0], fields[1]
                if ga != "." and gb != ".":
                    score = cscore_map.get((ga, gb), cscore_map.get((gb, ga), None))
                    genome_pairs.append((ga, gb, score))

        for i, (ga, gb, score) in enumerate(genome_pairs, 2):
            a_info = bed_a_idx.get(ga, ("", "", "", ""))
            b_info = bed_b_idx.get(gb, ("", "", "", ""))
            _write_row(ws2, i, [
                ga, a_info[0], a_info[1], a_info[2], a_info[3],
                gb, b_info[0], b_info[1], b_info[2], b_info[3],
                _format_score(score),
            ])

        for col_idx in range(1, len(all_headers) + 1):
            max_len = len(str(ws2.cell(row=1, column=col_idx).value))
            sample_count = min(len(genome_pairs), 200)
            for r in range(2, sample_count + 2):
                v = ws2.cell(row=r, column=col_idx).value
                if v:
                    max_len = max(max_len, len(str(v)))
            ws2.column_dimensions[ws2.cell(row=1, column=col_idx).column_letter].width = min(max_len + 3, 40)

        excel_path = micro_dir / "collinear_pairs.xlsx"
        wb.save(str(excel_path))

        self.logger.info(
            f"    区域内共线性配对 {len(all_pairs)} 对 "
            f"(目标相关: {row - 2 - len(unpaired_a) - len(unpaired_b)}, "
            f"无配对目标: {len(unpaired_a)}+{len(unpaired_b)})"
        )
        self.logger.info(f"    全基因组共线性配对 {len(genome_pairs)} 对")
        self.logger.info(f"    输出|Output: {excel_path}")

    def _build_cscore_index(self) -> dict:
        """从filtered.last构建C-score索引|Build C-score index from filtered.last"""
        cscore_map = {}
        stem_a, stem_b = get_jcvi_stem(self._current_name_a), get_jcvi_stem(self._current_name_b)
        last_file = (
            Path(self.config.output_dir) / "03_pairwise"
            / f"{self._current_name_a}_vs_{self._current_name_b}"
            / f"{stem_a}.{stem_b}.last.filtered"
        )
        if not last_file.exists():
            self.logger.warning(f"    未找到filtered.last, C-score将为空|filtered.last not found: {last_file}")
            return cscore_map

        with open(last_file, "r") as f:
            for line in f:
                parts = line.strip().split("\t")
                if len(parts) >= 3:
                    ga, gb = parts[0], parts[1]
                    try:
                        cscore_map[(ga, gb)] = float(parts[2])
                    except ValueError:
                        pass
        return cscore_map

    def _generate_layout(self, micro_dir: Path,
                         name_a: str, name_b: str) -> Path:
        """生成blocks.layout文件|Generate blocks.layout file"""
        layout_file = micro_dir / "blocks.layout"

        with open(layout_file, "w") as f:
            f.write(
                "# x,   y, rotation,   ha,     va,   color, ratio,            label\n"
            )
            f.write(
                f"0.5, 0.6,        0, left, center,       m,     1,       {name_a}\n"
            )
            f.write(
                f"0.5, 0.4,        0, left, center, #fc8d62,     1, {name_b}\n"
            )
            f.write("# edges\n")
            f.write("e, 0, 1\n")

        return layout_file

    def _plot_synteny(self, blocks_file: Path, merged_bed: Path,
                      layout_file: Path, micro_dir: Path,
                      suffix: str = "filtered",
                      genelabelsize: int = 6) -> Optional[Path]:
        """绘制microsynteny图|Plot microsynteny

        通过wrapper脚本覆盖JCVI的OrthoGroupPalette调色板,
        使所有共线性基因使用统一颜色, 非共线性singleton使用灰色。
        """
        # JCVI总是输出以blocks文件stem命名的PDF, 绘制后重命名为目标文件名
        jcvi_pdf = micro_dir / "filtered.pdf"
        wrapper_code = (
            "import sys\n"
            "from jcvi.graphics.synteny import OrthoGroupPalette\n"
            "OrthoGroupPalette.palette = ['#4477AA'] * len(OrthoGroupPalette.palette)\n"
            "from jcvi.graphics.synteny import main\n"
            "main(sys.argv[1:])\n"
        )

        cmd_args = [
            str(blocks_file.resolve()),
            str(merged_bed.resolve()),
            str(layout_file.resolve()),
            f"--genelabelsize={genelabelsize}",
            "--glyphcolor=orthogroup",
        ]

        if self.config.glyph_style:
            cmd_args.append(f"--glyphstyle={self.config.glyph_style}")
        if self.config.shadestyle:
            cmd_args.append(f"--shadestyle={self.config.shadestyle}")

        env_python = self._get_env_python()
        if env_python:
            # 写入临时wrapper脚本, 覆盖调色板后调用JCVI
            wrapper = micro_dir / "_plot_wrapper.py"
            wrapper.write_text(wrapper_code)
            try:
                cmd = [env_python, str(wrapper.resolve())] + cmd_args
                run_env = os.environ.copy()
                run_env["MPLBACKEND"] = "Agg"

                self.logger.info(f"    命令|Command: {' '.join(cmd)}")
                result = subprocess.run(cmd, cwd=str(micro_dir.resolve()), env=run_env,
                                        capture_output=True, text=True)
                if result.returncode != 0:
                    self.logger.error(f"    wrapper出错|wrapper failed: {result.stderr[:300]}")
            finally:
                wrapper.unlink(missing_ok=True)
        else:
            cmd = build_jcvi_command("jcvi.graphics.synteny", cmd_args,
                                     self.config.conda_env)

            self.logger.info(f"    命令|Command: {' '.join(cmd)}")
            run_env = None
            if cmd and "/envs/" in cmd[0] and cmd[0].endswith("python"):
                env_bin = os.path.dirname(cmd[0])
                run_env = os.environ.copy()
                run_env["PATH"] = env_bin + os.pathsep + run_env.get("PATH", "")
                run_env["MPLBACKEND"] = "Agg"
            result = subprocess.run(cmd, cwd=str(micro_dir), env=run_env)

        pdf_file = micro_dir / f"{suffix}.pdf"
        if jcvi_pdf.exists() and jcvi_pdf.stat().st_size > 0:
            jcvi_pdf.rename(pdf_file)
            return pdf_file
        self.logger.error("    synteny绘图失败|synteny plot failed")
        return None

    def _get_env_python(self) -> Optional[str]:
        """获取conda环境python绝对路径|Get conda env python absolute path"""
        from ..utils import _get_conda_env_python
        return _get_conda_env_python(self.config.conda_env)

    # ---- 转录本ID解析 ----

    def _find_gff_file(self, name: str) -> Optional[Path]:
        """根据样本名查找对应的GFF文件|Find GFF file for a sample name"""
        for ext in [self.config.gff_ext]:
            for candidate in [
                Path(self.config.input_dir) / f"{name}{ext}",
            ]:
                if candidate.exists():
                    return candidate
        return None

    def _build_transcript_map(self, name_a: str) -> Dict[str, Set[str]]:
        """从GFF构建转录本同源映射|Build transcript co-membership map from GFF

        返回 {transcript_id: {同一gene下的所有transcript_ids}}
        用于将用户传入的基因ID(如.m1)解析为BED中实际存在的ID(如.m2)
        """
        gff_file = self._find_gff_file(name_a)
        if not gff_file:
            self.logger.debug(f"    未找到GFF文件, 跳过转录本映射|GFF not found, skip transcript map: {name_a}")
            return {}

        gene_to_transcripts: Dict[str, Set[str]] = {}
        try:
            open_fn = open
            if str(gff_file).endswith(".gz"):
                import gzip
                open_fn = gzip.open

            with open_fn(gff_file, "rt") as f:
                for line in f:
                    if line.startswith("#"):
                        continue
                    parts = line.rstrip("\n").split("\t")
                    if len(parts) < 9:
                        continue
                    if parts[2] not in ("mRNA", "transcript"):
                        continue
                    attrs = parts[8]
                    transcript_id = None
                    gene_id = None
                    for attr in attrs.split(";"):
                        attr = attr.strip()
                        if attr.startswith("ID="):
                            transcript_id = attr[3:].strip()
                        elif attr.startswith("Parent="):
                            gene_id = attr[7:].strip()
                    if transcript_id and gene_id:
                        gene_to_transcripts.setdefault(gene_id, set()).add(transcript_id)
        except Exception as e:
            self.logger.warning(f"    解析GFF转录本映射失败|Failed to parse GFF transcript map: {e}")
            return {}

        transcript_map: Dict[str, Set[str]] = {}
        for gene_id, transcripts in gene_to_transcripts.items():
            if len(transcripts) > 1:
                for tid in transcripts:
                    transcript_map[tid] = transcripts

        mapped_count = sum(1 for t in transcript_map if len(transcript_map[t]) > 1)
        self.logger.info(f"    转录本映射|Transcript map ({name_a}): {mapped_count} 个基因有多转录本")
        return transcript_map

    def _resolve_gene_ids(
        self,
        gene_list: Set[str],
        bed_index: Dict[str, Tuple[str, int, int]],
        transcript_map: Dict[str, Set[str]],
    ) -> Set[str]:
        """将用户传入的基因ID解析为BED中实际存在的ID|Resolve user gene IDs to actual BED IDs

        1. 先检查用户ID是否直接存在于BED中
        2. 若不存在, 通过转录本同源映射找到兄弟转录本, 再检查BED
        """
        resolved = set()
        for gene_id in gene_list:
            if gene_id in bed_index:
                resolved.add(gene_id)
                continue
            # 尝试通过转录本映射找兄弟转录本
            siblings = transcript_map.get(gene_id, set())
            if not siblings:
                self.logger.debug(f"    基因 {gene_id} 在BED和GFF转录本映射中均未找到")
                continue
            found = False
            for sibling in siblings:
                if sibling in bed_index:
                    resolved.add(sibling)
                    if sibling != gene_id:
                        self.logger.debug(
                            f"    基因ID替换|Gene ID replaced: {gene_id} → {sibling}"
                        )
                    found = True
                    break
            if not found:
                self.logger.debug(f"    基因 {gene_id} 的所有转录本 {siblings} 均不在BED中")
        return resolved

    # ---- BED / Block 工具方法 ----

    def _load_bed_index(self, bed_file: Path) -> Dict[str, Tuple[str, int, int]]:
        """加载BED文件, 建立基因ID到坐标的索引|Load BED, build gene->coordinate index"""
        index = {}
        with open(bed_file, "r") as f:
            for line in f:
                line = line.rstrip("\n")
                if not line or line.startswith("#"):
                    continue
                fields = line.split("\t")
                if len(fields) >= 4:
                    chrom, start, end, gene_id = (
                        fields[0], int(fields[1]), int(fields[2]), fields[3]
                    )
                    index[gene_id] = (chrom, start, end)
        return index

    def _parse_region(self, region_str: str) -> Optional[Tuple[str, int, int]]:
        """解析区域字符串 chr:start-end|Parse region string chr:start-end"""
        try:
            parts = region_str.split(":")
            if len(parts) != 2:
                self.logger.error(
                    f"区域格式错误|Invalid region format: {region_str}"
                )
                return None
            chrom = parts[0]
            pos_parts = parts[1].split("-")
            if len(pos_parts) != 2:
                self.logger.error(
                    f"区域格式错误|Invalid region format: {region_str}"
                )
                return None
            start = int(pos_parts[0])
            end = int(pos_parts[1])
            return (chrom, start, end)
        except (ValueError, IndexError) as e:
            self.logger.error(
                f"区域解析失败|Region parse failed: {region_str} ({e})"
            )
            return None

    def _find_region_genes(
        self,
        bed_index: Dict[str, Tuple[str, int, int]],
        region: Tuple[str, int, int],
    ) -> Set[str]:
        """查找坐标落在区域内的基因(返回完整BED ID)|Find genes within region (return full BED IDs)"""
        chrom, start, end = region
        genes = set()
        for gene_id, (g_chrom, g_start, g_end) in bed_index.items():
            if g_chrom == chrom and g_end > start and g_start < end:
                genes.add(gene_id)
        return genes

    def _parse_blocks(self, blocks_file: Path) -> List[List[str]]:
        """解析blocks文件为block列表|Parse blocks file into list of blocks"""
        blocks = []
        current = []

        with open(blocks_file, "r") as f:
            for line in f:
                stripped = line.rstrip("\n")
                if stripped == "###":
                    if current:
                        blocks.append(current)
                        current = []
                elif stripped.startswith("#"):
                    continue
                elif stripped:
                    current.append(stripped)

        if current:
            blocks.append(current)

        return blocks

    def _select_region_blocks(
        self,
        blocks: List[List[str]],
        genes_in_region_a: Set[str],
        genes_in_region_b: Set[str],
    ) -> List[int]:
        """筛选与两个区域都重叠的block索引|Select block indices overlapping both regions"""
        matching = []
        for idx, block in enumerate(blocks):
            has_a = False
            has_b = False
            for line in block:
                fields = line.split("\t")
                col0 = fields[0] if fields else ""
                col1 = fields[1] if len(fields) > 1 else ""
                if not has_a and col0 in genes_in_region_a:
                    has_a = True
                if col1 and not has_b and col1 in genes_in_region_b:
                    has_b = True
                if has_a and has_b:
                    break
            if has_a and has_b:
                matching.append(idx)
        return matching

    def _filter_gene_rows(
        self,
        block_lines: List[str],
        genes_a_set: Optional[Set[str]],
        genes_b_set: Optional[Set[str]],
    ) -> List[str]:
        """按基因列表过滤block行(用完整BED ID匹配)|Filter block rows by gene lists (full BED ID match)"""
        if genes_a_set is None and genes_b_set is None:
            return block_lines

        filtered = []
        for line in block_lines:
            fields = line.split("\t")
            col0 = fields[0] if fields else ""
            col1 = fields[1] if len(fields) > 1 else ""

            if genes_a_set is not None and genes_b_set is None:
                if col0 in genes_a_set:
                    filtered.append(line)
            elif genes_a_set is None and genes_b_set is not None:
                if col1 in genes_b_set:
                    filtered.append(line)
            else:
                if col0 in genes_a_set or col1 in genes_b_set:
                    filtered.append(line)

        return filtered

    @staticmethod
    def _parse_gene_list(gene_file: Optional[str]) -> Optional[Set[str]]:
        """从文件读取基因列表, 一行一个|Read gene list from file, one per line"""
        if not gene_file:
            return None
        genes = set()
        with open(gene_file, "r") as f:
            for line in f:
                gene_id = line.strip()
                if gene_id:
                    genes.add(gene_id)
        return genes if genes else None

    # ---- 通用工具 ----

    def _find_bed(self, name: str) -> Optional[Path]:
        """查找BED文件(同时检查完整名和JCVI stem名)|Find BED file"""
        stem = get_jcvi_stem(name)
        for path in [
            Path(self.config.output_dir) / "02_bed" / f"{name}.uniq.bed",
            Path(self.config.output_dir) / "02_bed" / f"{stem}.uniq.bed",
            Path(self.config.output_dir) / "03_pairwise" / f"{name}.bed",
            Path(self.config.output_dir) / "03_pairwise" / f"{stem}.bed",
        ]:
            if path.exists():
                return path
        return None

    def _find_pair_dir(self, name_a: str, name_b: str) -> Optional[Path]:
        """查找pair目录|Find pair directory"""
        pair_dir = (
            Path(self.config.output_dir) / "03_pairwise" / f"{name_a}_vs_{name_b}"
        )
        if pair_dir.exists():
            return pair_dir
        return None

    def _run_cmd(self, cmd, cwd=None):
        """执行命令|Execute command"""
        run_env = None
        if cmd and "/envs/" in cmd[0] and cmd[0].endswith("python"):
            env_bin = os.path.dirname(cmd[0])
            run_env = os.environ.copy()
            run_env["PATH"] = env_bin + os.pathsep + run_env.get("PATH", "")
        result = subprocess.run(cmd, cwd=cwd, env=run_env)
        return result.returncode

    def _print_header(self):
        if not self.logger:
            return
        self.logger.info("")
        self.logger.info("=" * 60)
        self.logger.info("JCVI 微观共线性可视化|JCVI Microsynteny Visualization")
        self.logger.info("=" * 60)
        self.logger.info(f"版本|Version: {self.VERSION}")
        self.logger.info(f"时间|Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info("")
        if self.config.pairs:
            pairs_str = (
                self.config.pairs[0]
                if isinstance(self.config.pairs, list)
                else self.config.pairs
            )
            self.logger.info(f"  物种对|Species pair: {pairs_str}")
        self.logger.info(f"  区域A|Region A: {self.config.region_a}")
        self.logger.info(f"  区域B|Region B: {self.config.region_b}")
        if self.config.genes_a:
            self.logger.info(f"  展示基因A|Genes A: {self.config.genes_a}")
        if self.config.genes_b:
            self.logger.info(f"  展示基因B|Genes B: {self.config.genes_b}")
        self.logger.info("")

    def _print_footer(self, success: bool):
        if not self.logger:
            return
        elapsed = time.time() - self.start_time
        hours, remainder = divmod(int(elapsed), 3600)
        minutes, seconds = divmod(remainder, 60)
        self.logger.info("")
        self.logger.info("=" * 60)
        self.logger.info(
            "微观共线性可视化完成|Microsynteny Visualization Completed"
            if success
            else "微观共线性可视化失败|Failed"
        )
        self.logger.info(f"  总耗时|Total time: {hours}h {minutes}m {seconds}s")
        self.logger.info("=" * 60)
