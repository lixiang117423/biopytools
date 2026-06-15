"""ATOMM工具函数模块|ATOMM Utility Functions Module"""

import logging
import sys
import numpy as np


class ATOMMLogger:
    """ATOMM日志管理器|ATOMM Logger Manager"""

    def __init__(self, log_file=None, log_level="INFO"):
        self.log_file = log_file
        self.setup_logging(log_level)

    def setup_logging(self, log_level):
        """设置日志|Setup logging"""
        log_format = '%(asctime)s.%(msecs)03d - %(levelname)s - %(message)s'
        date_format = '%Y-%m-%d %H:%M:%S'

        level = getattr(logging, log_level.upper(), logging.INFO)
        handlers = [logging.StreamHandler(sys.stdout)]
        if self.log_file:
            handlers.append(logging.FileHandler(self.log_file))

        logging.basicConfig(
            level=level,
            format=log_format,
            datefmt=date_format,
            handlers=handlers,
            force=True
        )
        self.logger = logging.getLogger(__name__)

    def get_logger(self):
        """获取日志器|Get logger"""
        return self.logger


def read_genotype(filepath):
    """读取基因型文件|Read genotype file

    Args:
        filepath: 基因型文件路径|Genotype file path

    Returns:
        tuple: (chrom_ids, snp_ids, genotype_matrix)
            chrom_ids: 染色体ID列表|Chromosome ID list
            snp_ids: SNP ID列表|SNP ID list
            genotype_matrix: 基因型矩阵 (n_snps x n_individuals), 0/1编码
    """
    chrom_ids = []
    snp_ids = []
    genotypes = []

    with open(filepath, 'r') as f:
        for line in f:
            parts = line.strip().split()
            chrom_ids.append(parts[0])
            snp_ids.append(int(parts[1]))
            genotypes.append([float(x) for x in parts[2:]])

    return chrom_ids, snp_ids, np.array(genotypes)


def read_phenotype(filepath):
    """读取表型文件|Read phenotype file

    Args:
        filepath: 表型文件路径|Phenotype file path

    Returns:
        tuple: (host_ids, pathogen_ids, covariates, phenotypes)
            host_ids: 宿主ID数组 (n_obs,)|Host ID array
            pathogen_ids: 病原ID数组 (n_obs,)|Pathogen ID array
            covariates: 协变量矩阵 (n_obs, n_cov)|Covariate matrix (includes intercept)
            phenotypes: 表型值数组 (n_obs,)|Phenotype array
    """
    data = np.loadtxt(filepath)
    host_ids = data[:, 0].astype(int)
    pathogen_ids = data[:, 1].astype(int)
    phenotypes = data[:, -1]
    covariates = data[:, 2:-1]
    return host_ids, pathogen_ids, covariates, phenotypes


def _safe_neglog10(p):
    """安全计算-log10(p)|Safely compute -log10(p)"""
    if p <= 0:
        return 300.0
    return -np.log10(p)


def _write_marginal_tsv(filepath, chrom_ids, snp_ids, statistics, p_values, header_label):
    """写入边际检验TSV结果|Write marginal test TSV results

    Args:
        filepath: 输出文件路径|Output file path
        chrom_ids: 染色体ID列表|Chromosome ID list
        snp_ids: SNP ID列表|SNP ID list
        statistics: 检验统计量|Test statistics
        p_values: p值|P-values
        header_label: 列标签前缀|Column label prefix
    """
    with open(filepath, 'w') as f:
        f.write(f"{header_label}_Chr\t{header_label}_POS\tStatistic\tPvalue\tNegLog10P\n")
        for i in range(len(chrom_ids)):
            neg_log_p = _safe_neglog10(p_values[i])
            f.write(f"{chrom_ids[i]}\t{snp_ids[i]}\t{statistics[i]:.4f}\t"
                    f"{p_values[i]:.6e}\t{neg_log_p:.4f}\n")


def write_estimate_tsv(filepath, herit):
    """写入遗传力估计TSV结果|Write heritability estimates TSV

    Args:
        filepath: 输出文件路径|Output file path
        herit: 遗传力数组 [xi_h, xi_p, xi_hp, xi_e]
    """
    labels = ['Host', 'Pathogen', 'Interaction', 'Residual']
    with open(filepath, 'w') as f:
        f.write("Component\tEstimate\n")
        for label, val in zip(labels, herit):
            f.write(f"{label}\t{val:.6f}\n")


def write_marginal_host_tsv(filepath, chrom_ids, snp_ids, statistics, p_values):
    """写入宿主边际检验TSV结果|Write host marginal test TSV results"""
    _write_marginal_tsv(filepath, chrom_ids, snp_ids, statistics, p_values, "Host")


def write_marginal_pathogen_tsv(filepath, chrom_ids, snp_ids, statistics, p_values):
    """写入病原边际检验TSV结果|Write pathogen marginal test TSV results"""
    _write_marginal_tsv(filepath, chrom_ids, snp_ids, statistics, p_values, "Pathogen")


def write_interaction_tsv(filepath, host_chrom, host_snp, pathogen_chrom, pathogen_snp,
                           statistics, p_values):
    """写入交互检验TSV结果|Write interaction test TSV results"""
    with open(filepath, 'w') as f:
        f.write("Host_Chr\tHost_POS\tPathogen_Chr\tPathogen_POS\tStatistic\tPvalue\tNegLog10P\n")
        for i in range(len(host_chrom)):
            neg_log_p = _safe_neglog10(p_values[i])
            f.write(f"{host_chrom[i]}\t{host_snp[i]}\t{pathogen_chrom[i]}\t"
                    f"{pathogen_snp[i]}\t{statistics[i]:.4f}\t{p_values[i]:.6e}\t"
                    f"{neg_log_p:.4f}\n")


def write_results_excel(filepath, herit=None, host_data=None, pathogen_data=None,
                        interaction_data=None):
    """写入Excel汇总结果|Write summary results to Excel

    Args:
        filepath: 输出文件路径(.xlsx)|Output file path (.xlsx)
        herit: 遗传力数组 [xi_h, xi_p, xi_hp, xi_e]|Heritability array
        host_data: 宿主边际检验元组 (chrom_ids, snp_ids, statistics, p_values)|Host marginal data
        pathogen_data: 病原边际检验元组|Pathogen marginal data
        interaction_data: 交互检验元组 (h_chr, h_snp, p_chr, p_snp, stat, pval)|Interaction data
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        import logging
        logging.getLogger(__name__).warning(
            "openpyxl未安装，跳过Excel输出|openpyxl not installed, skipping Excel output"
        )
        return

    wb = openpyxl.Workbook()

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")
    data_align = Alignment(horizontal="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    def _style_header(ws, ncols):
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border

    def _style_data(ws, nrows, ncols):
        for row in range(2, nrows + 2):
            for col in range(1, ncols + 1):
                cell = ws.cell(row=row, column=col)
                cell.alignment = data_align
                cell.border = thin_border

    def _auto_width(ws, ncols):
        for col in range(1, ncols + 1):
            max_len = 0
            for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
                for cell in row:
                    val = str(cell.value) if cell.value else ""
                    max_len = max(max_len, len(val))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = min(max_len + 4, 30)

    # Sheet 1: Heritability
    if herit is not None:
        ws = wb.active
        ws.title = "Heritability"
        labels = ['Host_xi_h', 'Pathogen_xi_p', 'Interaction_xi_hp', 'Residual_xi_e']
        ws.append(["Component", "Estimate", "Proportion(%)"])
        total = sum(herit)
        for label, val in zip(labels, herit):
            pct = val / total * 100 if total > 0 else 0
            ws.append([label, round(val, 6), round(pct, 2)])
        _style_header(ws, 3)
        _style_data(ws, len(labels), 3)
        _auto_width(ws, 3)

    # Sheet 2: Host Marginal
    if host_data is not None:
        chrom_ids, snp_ids, statistics, p_values = host_data
        ws = wb.create_sheet("Host_Marginal")
        ws.append(["Host_Chr", "Host_POS", "Statistic", "Pvalue", "-log10(P)"])
        for i in range(len(chrom_ids)):
            neg_log_p = _safe_neglog10(p_values[i])
            ws.append([
                str(chrom_ids[i]), int(snp_ids[i]),
                round(float(statistics[i]), 4),
                float(p_values[i]),
                round(neg_log_p, 4)
            ])
        _style_header(ws, 5)
        _style_data(ws, len(chrom_ids), 5)
        _auto_width(ws, 5)

    # Sheet 3: Pathogen Marginal
    if pathogen_data is not None:
        chrom_ids, snp_ids, statistics, p_values = pathogen_data
        ws = wb.create_sheet("Pathogen_Marginal")
        ws.append(["Pathogen_Chr", "Pathogen_POS", "Statistic", "Pvalue", "-log10(P)"])
        for i in range(len(chrom_ids)):
            neg_log_p = _safe_neglog10(p_values[i])
            ws.append([
                str(chrom_ids[i]), int(snp_ids[i]),
                round(float(statistics[i]), 4),
                float(p_values[i]),
                round(neg_log_p, 4)
            ])
        _style_header(ws, 5)
        _style_data(ws, len(chrom_ids), 5)
        _auto_width(ws, 5)

    # Sheet 4: Interaction
    if interaction_data is not None:
        h_chr, h_snp, p_chr, p_snp, statistics, p_values = interaction_data
        ws = wb.create_sheet("Interaction")
        ws.append(["Host_Chr", "Host_POS", "Pathogen_Chr", "Pathogen_POS",
                    "Statistic", "Pvalue", "-log10(P)"])
        for i in range(len(h_chr)):
            neg_log_p = _safe_neglog10(p_values[i])
            ws.append([
                str(h_chr[i]), int(h_snp[i]),
                str(p_chr[i]), int(p_snp[i]),
                round(float(statistics[i]), 4),
                float(p_values[i]),
                round(neg_log_p, 4)
            ])
        _style_header(ws, 7)
        _style_data(ws, len(h_chr), 7)
        _auto_width(ws, 7)

    wb.save(filepath)


def format_number(num):
    """格式化数字|Format number"""
    if num >= 1_000_000:
        return f"{num / 1_000_000:.2f}M"
    if num >= 1_000:
        return f"{num / 1_000:.1f}K"
    return str(int(num))
