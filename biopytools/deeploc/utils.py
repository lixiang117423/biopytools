"""
DeepLoc 2.1蛋白质亚细胞定位预测工具函数模块|DeepLoc 2.1 Protein Subcellular Localization Prediction Utility Functions Module
"""

import logging
import os
import subprocess
import sys
import re
import shutil
from pathlib import Path
from typing import Optional, List


def get_conda_env(command: str) -> Optional[str]:
    """
    检测命令是否在conda环境中，返回环境名称|Detect if command is in conda environment, return env name

    Args:
        command: 命令名称或路径|Command name or path

    Returns:
        conda环境名称或None|conda environment name or None
    """
    # 首先尝试从命令路径检测|First try to detect from command path
    cmd_path = shutil.which(command)
    if cmd_path:
        # 检查路径中是否包含 envs|Check if path contains 'envs'
        match = re.search(r'/envs/([^/]+)', cmd_path)
        if match:
            return match.group(1)

    # 如果未找到，尝试搜索conda环境|If not found, try searching conda environments
    conda_base = os.environ.get('CONDA_EXE')
    if conda_base:
        conda_base_dir = os.path.dirname(os.path.dirname(conda_base))
        envs_dir = os.path.join(conda_base_dir, 'envs')

        if os.path.exists(envs_dir):
            for env_name in os.listdir(envs_dir):
                env_bin = os.path.join(envs_dir, env_name, 'bin', command)
                if os.path.exists(env_bin):
                    return env_name

    return None


def build_conda_command(command: str, args: List[str]) -> List[str]:
    """
    构建conda run命令来运行conda环境中的软件|Build conda run command to run software in conda environment

    Args:
        command: 命令名称|Command name
        args: 命令参数|Command arguments

    Returns:
        完整命令列表|Complete command list
    """
    # 检查是否在conda环境中|Check if in conda environment
    conda_env = get_conda_env(command)

    if conda_env:
        # 使用 conda run|Use conda run
        full_cmd = ['conda', 'run', '-n', conda_env, '--no-capture-output', command] + args
    else:
        # 直接调用|Direct call
        full_cmd = [command] + args

    return full_cmd


class DeepLocLogger:
    """DeepLoc亚细胞定位预测日志管理器|DeepLoc Subcellular Localization Prediction Logger Manager"""

    def __init__(self, output_dir: Path, log_name: str = "deeploc_prediction.log"):
        self.output_dir = output_dir
        self.log_file = output_dir / log_name
        self.setup_logging()

    def setup_logging(self):
        """设置日志|Setup logging"""
        # 删除旧日志文件|Delete old log file
        if self.log_file.exists():
            self.log_file.unlink()

        # 设置日志格式|Set log format
        formatter = logging.Formatter(
            '%(asctime)s.%(msecs)03d - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )

        # 文件handler|File handler
        file_handler = logging.FileHandler(self.log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        file_handler.setFormatter(formatter)

        # stdout handler|Stdout handler
        stdout_handler = logging.StreamHandler(sys.stdout)
        stdout_handler.setLevel(logging.INFO)
        stdout_handler.setFormatter(formatter)

        # 配置日志|Configure logging
        logging.basicConfig(
            level=logging.DEBUG,
            handlers=[file_handler, stdout_handler]
        )
        self.logger = logging.getLogger(__name__)

    def get_logger(self):
        """获取日志器|Get logger"""
        return self.logger


class DeepLocRunner:
    """DeepLoc命令执行器|DeepLoc Command Runner"""

    def __init__(self, config, logger):
        self.config = config
        self.logger = logger
        self.csv_result = None

    def build_command(self):
        """构建DeepLoc命令|Build DeepLoc command"""
        # 构建singularity exec命令|Build singularity exec command
        cmd_parts = [
            self.config.singularity_exec,
            "exec",
            # 挂载数据库目录|Mount database directory
            "-B", f"{self.config.database_dir}:/database",
            self.config.singularity_image,
            "deeploc2",
            "-f", "/input/" + os.path.basename(self.config.fasta_file),  # 在容器内的路径
            "-o", "/output",  # 容器内的输出目录
            "-m", self.config.model,
            "-d", self.config.device
        ]

        # 可选参数|Optional parameters
        if self.config.plot:
            cmd_parts.append("-p")

        return cmd_parts

    def run(self):
        """运行DeepLoc预测|Run DeepLoc prediction"""
        self.logger.info("=" * 80)
        self.logger.info("开始DeepLoc 2.1亚细胞定位预测|Starting DeepLoc 2.1 Subcellular Localization Prediction")
        self.logger.info("=" * 80)

        # 显示配置信息|Show configuration
        self.logger.info(f"输入文件|Input file: {self.config.fasta_file}")
        self.logger.info(f"输出目录|Output directory: {self.config.output_dir}")
        self.logger.info(f"预测模型|Model: {self.config.model}")
        self.logger.info(f"计算设备|Device: {self.config.device}")
        self.logger.info(f"绘制Attention图|Plot attention: {self.config.plot}")

        self.logger.info(f"Singularity镜像|Singularity image: {self.config.singularity_image}")
        self.logger.info(f"数据库目录|Database directory: {self.config.database_dir}")

        self.logger.info("=" * 80)

        # 构建命令|Build command
        # 注意：需要将输入文件挂载到容器中
        input_dir = os.path.dirname(self.config.fasta_file)
        input_filename = os.path.basename(self.config.fasta_file)

        cmd = [
            self.config.singularity_exec,
            "exec",
            # 挂载输入文件目录|Mount input file directory
            "-B", f"{input_dir}:/input",
            # 挂载输出目录|Mount output directory
            "-B", f"{self.config.output_dir}:/output",
            # 挂载数据库目录|Mount database directory
            "-B", f"{self.config.database_dir}:/database",
            self.config.singularity_image,
            "deeploc2",
            "-f", f"/input/{input_filename}",
            "-o", "/output",
            "-m", self.config.model,
            "-d", self.config.device
        ]

        # 添加可选参数|Add optional parameters
        if self.config.plot:
            cmd.append("-p")

        # 使用conda wrapper包装singularity命令|Use conda wrapper to wrap singularity command
        singularity_cmd = os.path.basename(self.config.singularity_exec)
        cmd_args = cmd[1:]  # 跳过第一个元素（singularity路径）|Skip first element (singularity path)
        wrapped_cmd = build_conda_command(singularity_cmd, cmd_args)

        cmd_str = " ".join(wrapped_cmd)
        self.logger.info(f"执行命令|Executing command: {cmd_str}")
        self.logger.info("")

        try:
            # 执行命令|Execute command
            result = subprocess.run(
                wrapped_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                check=False
            )

            # 输出结果|Output results
            self.logger.info(result.stdout)

            if result.returncode != 0:
                self.logger.error(f"DeepLoc执行失败|DeepLoc execution failed (exit code: {result.returncode})")
                return False

            self.logger.info("")
            self.logger.info("=" * 80)
            self.logger.info("DeepLoc预测完成|DeepLoc prediction completed")
            self.logger.info("=" * 80)

            # 检查输出文件|Check output files
            self._check_output_files()

            # 格式化结果|Format results
            self._format_results()

            return True

        except Exception as e:
            self.logger.error(f"预测出错|Prediction error: {e}")
            return False

    def _check_output_files(self):
        """检查输出文件|Check output files"""
        self.logger.info("检查输出文件|Checking output files:")

        # DeepLoc2输出的主要文件是命名基于输入FASTA文件名的
        # 例如：input.fasta -> input_deeploc2.tsv
        input_basename = os.path.splitext(os.path.basename(self.config.fasta_file))[0]
        expected_output = f"{input_basename}_deeploc2.tsv"
        expected_plot = f"{input_basename}_deeploc2_plot.png"

        # 检查预测结果文件|Check prediction result file
        output_file = os.path.join(self.config.output_dir, expected_output)
        if os.path.exists(output_file):
            size = os.path.getsize(output_file)
            self.logger.info(f"已生成|Generated: {expected_output} ({size} bytes)")
        else:
            self.logger.warning(f"未生成|Not generated: {expected_output}")

        # 检查绘图文件（如果启用了-p）|Check plot file (if -p enabled)
        if self.config.plot:
            plot_file = os.path.join(self.config.output_dir, expected_plot)
            if os.path.exists(plot_file):
                size = os.path.getsize(plot_file)
                self.logger.info(f"已生成|Generated: {expected_plot} ({size} bytes)")
            else:
                self.logger.warning(f"未生成|Not generated: {expected_plot}")

        # 检查CSV结果文件|Check CSV result file
        import glob
        csv_files = glob.glob(os.path.join(self.config.output_dir, "results_*.csv"))
        if csv_files:
            self.csv_result = csv_files[0]
            size = os.path.getsize(self.csv_result)
            self.logger.info(f"已生成|Generated: {os.path.basename(self.csv_result)} ({size} bytes)")
        else:
            self.csv_result = None
            self.logger.warning("results_*.csv 未生成|results_*.csv not generated")

    def _format_results(self):
        """格式化预测结果|Format prediction results"""
        if not self.csv_result or not os.path.exists(self.csv_result):
            self.logger.warning("未找到CSV结果文件，跳过格式化|CSV result file not found, skip formatting")
            return

        formatter = DeepLocResultFormatter(
            csv_file=self.csv_result,
            output_dir=self.config.output_dir,
            logger=self.logger
        )

        # 生成TSV格式|Generate TSV format
        formatter.format_tsv()

        # 生成Excel格式|Generate Excel format
        formatter.format_excel()


def format_number(num: int) -> str:
    """格式化数字|Format number

    大于1百万的数字使用M单位|Numbers larger than 1M use M unit
    """
    if num >= 1_000_000:
        return f"{num / 1_000_000:.2f}M"
    elif num >= 1_000:
        return f"{num / 1_000:.2f}K"
    return str(num)


class DeepLocResultFormatter:
    """DeepLoc结果格式化器|DeepLoc Result Formatter

    将DeepLoc CSV输出转换为Excel和TSV格式
    Convert DeepLoc CSV output to Excel and TSV formats
    """

    # 亚细胞定位中文名称|Subcellular localization Chinese names
    LOCALIZATION_ZH = {
        'Cytoplasm': '细胞质',
        'Nucleus': '细胞核',
        'Extracellular': '细胞外',
        'Cell membrane': '细胞膜',
        'Mitochondrion': '线粒体',
        'Plastid': '质体',
        'Endoplasmic reticulum': '内质网',
        'Lysosome/Vacuole': '溶酶体/液泡',
        'Golgi apparatus': '高尔基体',
        'Peroxisome': '过氧化物酶体'
    }

    # 信号类型中文名称|Signal type Chinese names
    SIGNAL_ZH = {
        'Signal peptide': '信号肽',
        'Transmembrane domain': '跨膜结构域',
        'Chloroplast transit peptide': '叶绿体转运肽',
        'Nuclear localization signal': '核定位信号',
        'Nuclear export signal': '核输出信号',
        'Mitochondrial transit peptide': '线粒体转运肽',
        'Peroxisomal targeting signal': '过氧化物酶体靶向信号'
    }

    # 膜蛋白类型中文名称|Membrane protein type Chinese names
    MEMBRANE_ZH = {
        'Peripheral': '外周膜',
        'Transmembrane': '跨膜',
        'Lipid anchor': '脂质锚定',
        'Soluble': '可溶性'
    }

    def __init__(self, csv_file: str, output_dir: str, logger=None):
        """初始化格式化器|Initialize formatter

        Args:
            csv_file: DeepLoc输出的CSV文件|DeepLoc output CSV file
            output_dir: 输出目录|Output directory
            logger: 日志器|Logger instance
        """
        self.csv_file = csv_file
        self.output_dir = output_dir
        self.logger = logger

    def format_probability(self, prob_str: str) -> str:
        """格式化概率为百分比|Format probability as percentage"""
        try:
            prob = float(prob_str)
            return f"{prob * 100:.2f}%"
        except (ValueError, TypeError):
            return "N/A"

    def format_localization(self, loc_str: str) -> str:
        """将定位翻译为中文|Translate localization to Chinese"""
        if not loc_str:
            return "-"

        locs = loc_str.split('|')
        zh_locs = []
        for loc in locs:
            zh_loc = self.LOCALIZATION_ZH.get(loc, loc)
            zh_locs.append(zh_loc)

        return "|".join(zh_locs)

    def format_signals(self, signal_str: str) -> str:
        """将信号翻译为中文|Translate signals to Chinese"""
        if not signal_str:
            return "-"

        signals = signal_str.split('|')
        zh_signals = []
        for signal in signals:
            zh_signal = self.SIGNAL_ZH.get(signal, signal)
            zh_signals.append(zh_signal)

        return "|".join(zh_signals)

    def format_membrane(self, membrane_str: str) -> str:
        """将膜类型翻译为中文|Translate membrane type to Chinese"""
        return self.MEMBRANE_ZH.get(membrane_str, membrane_str)

    def format_tsv(self):
        """生成TSV格式|Generate TSV format"""
        import csv

        tsv_file = os.path.join(self.output_dir, "deeploc_results_readable.tsv")

        if self.logger:
            self.logger.info("生成TSV格式|Generating TSV format")

        try:
            with open(self.csv_file, 'r', encoding='utf-8') as f_in, \
                 open(tsv_file, 'w', encoding='utf-8') as f_out:

                reader = csv.reader(f_in)
                headers = next(reader)

                # 写入中文表头|Write Chinese headers
                zh_headers = [
                    "蛋白质ID|Protein ID",
                    "亚细胞定位|Localization",
                    "信号序列|Signals",
                    "膜蛋白类型|Membrane Type",
                    "细胞质|Cytoplasm",
                    "细胞核|Nucleus",
                    "细胞外|Extracellular",
                    "细胞膜|Cell Membrane",
                    "线粒体|Mitochondrion",
                    "质体|Plastid",
                    "内质网|Endoplasmic Reticulum",
                    "溶酶体/液泡|Lysosome/Vacuole",
                    "高尔基体|Golgi Apparatus",
                    "过氧化物酶体|Peroxisome",
                    "外周膜|Peripheral",
                    "跨膜|Transmembrane",
                    "脂质锚定|Lipid Anchor",
                    "可溶性|Soluble"
                ]

                f_out.write("\t".join(zh_headers) + "\n")

                # 处理数据行|Process data rows
                count = 0
                for row in reader:
                    if len(row) < 18:
                        continue

                    # 格式化前几列为中文|Format first few columns to Chinese
                    protein_id = row[0]
                    localizations = self.format_localization(row[1])
                    signals = self.format_signals(row[2])
                    membrane = self.format_membrane(row[3])

                    # 格式化概率为百分比|Format probabilities as percentages
                    probs = [self.format_probability(p) for p in row[4:18]]

                    formatted_row = [protein_id, localizations, signals, membrane] + probs
                    f_out.write("\t".join(formatted_row) + "\n")
                    count += 1

            if self.logger:
                self.logger.info(f"TSV格式已保存|TSV format saved: {tsv_file}")
                self.logger.info(f"  蛋白质数量|Proteins: {count}")

            return True

        except Exception as e:
            if self.logger:
                self.logger.error(f"TSV格式化失败|TSV formatting failed: {e}")
            return False

    def format_excel(self):
        """生成Excel格式|Generate Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            if self.logger:
                self.logger.warning("未安装openpyxl，跳过Excel生成|openpyxl not installed, skip Excel generation")
                self.logger.warning("  安装命令|Install command: pip install openpyxl")
            return False

        import csv

        excel_file = os.path.join(self.output_dir, "deeploc_results.xlsx")

        if self.logger:
            self.logger.info("生成Excel格式|Generating Excel format")

        try:
            # 创建工作簿|Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "DeepLoc预测结果|Prediction Results"

            # 读取CSV数据|Read CSV data
            with open(self.csv_file, 'r', encoding='utf-8') as f_in:
                reader = csv.reader(f_in)
                headers = next(reader)

                # 写入表头|Write headers (中英文对照)
                header_row = [
                    "蛋白质ID\nProtein ID",
                    "亚细胞定位\nLocalization",
                    "信号序列\nSignals",
                    "膜蛋白类型\nMembrane Type",
                    "细胞质\nCytoplasm",
                    "细胞核\nNucleus",
                    "细胞外\nExtracellular",
                    "细胞膜\nCell Membrane",
                    "线粒体\nMitochondrion",
                    "质体\nPlastid",
                    "内质网\nEndoplasmic Reticulum",
                    "溶酶体/液泡\nLysosome/Vacuole",
                    "高尔基体\nGolgi Apparatus",
                    "过氧化物酶体\nPeroxisome",
                    "外周膜\nPeripheral",
                    "跨膜\nTransmembrane",
                    "脂质锚定\nLipid Anchor",
                    "可溶性\nSoluble"
                ]

                # 写入表头并设置样式|Write headers with style
                header_font = Font(bold=True, size=11)
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')

                for col_idx, header in enumerate(header_row, start=1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.fill = header_fill

                # 设置列宽|Set column widths
                ws.column_dimensions['A'].width = 15  # 蛋白质ID
                ws.column_dimensions['B'].width = 30  # 定位
                ws.column_dimensions['C'].width = 30  # 信号
                ws.column_dimensions['D'].width = 15  # 膜类型
                for col in range(5, 19):  # 概率列
                    ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 12

                # 写入数据|Write data
                count = 0
                row_num = 2
                for row in reader:
                    if len(row) < 18:
                        continue

                    # 格式化前几列为中文|Format first few columns to Chinese
                    protein_id = row[0]
                    localizations = self.format_localization(row[1])
                    signals = self.format_signals(row[2])
                    membrane = self.format_membrane(row[3])

                    # 概率值保留为小数（方便Excel计算）|Keep probabilities as decimals for Excel
                    data_row = [protein_id, localizations, signals, membrane] + row[4:18]

                    for col_idx, value in enumerate(data_row, start=1):
                        cell = ws.cell(row=row_num, column=col_idx, value=value)
                        # 居中对齐|Center align
                        cell.alignment = Alignment(horizontal='center', vertical='center')

                    row_num += 1
                    count += 1

            # 保存Excel文件|Save Excel file
            wb.save(excel_file)

            if self.logger:
                self.logger.info(f"Excel格式已保存|Excel format saved: {excel_file}")
                self.logger.info(f"  蛋白质数量|Proteins: {count}")

            return True

        except Exception as e:
            if self.logger:
                self.logger.error(f"Excel格式化失败|Excel formatting failed: {e}")
            return False
