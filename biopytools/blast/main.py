"""
BLAST分析主程序模块|BLAST Analysis Main Module
标准化BLAST分析器，遵循开发规范
"""

import os
import sys
import subprocess
import glob
import re
import shutil
from pathlib import Path
from typing import List, Dict, Optional, Tuple

from ..core.base_analyzer import BaseAnalyzer
from .config import BLASTConfig
from .utils import build_conda_command


class BLASTAnalyzer(BaseAnalyzer):
    """BLAST分析主类|Main BLAST Analysis Class"""

    def __init__(self, config: BLASTConfig):
        """
        初始化BLAST分析器|Initialize BLAST analyzer

        Args:
            config: BLAST配置对象|BLAST configuration object
        """
        super().__init__(config)
        self.sample_mapping = {}
        self.database_path = None

    def validate_inputs(self) -> bool:
        """
        验证输入|Validate inputs

        Returns:
            bool: 验证是否通过|Whether validation passed
        """
        try:
            # 调用父类验证
            super().validate_inputs()

            # 检查依赖工具|Check dependencies
            self._check_dependencies()

            # 生成或验证样品映射|Generate or validate sample mapping
            self._generate_sample_mapping()

            return True

        except Exception as e:
            self.logger.error(f"输入验证失败|Input validation failed: {e}")
            return False

    def _check_dependencies(self):
        """检查BLAST依赖工具|Check BLAST dependencies"""
        # 获取正确的BLAST程序路径|Get correct BLAST program path
        blast_program = getattr(self.config, f'{self.config.blast_type}_path', self.config.blast_type)

        required_tools = [self.config.makeblastdb_path, blast_program]

        for tool in required_tools:
            if not shutil.which(tool):
                raise RuntimeError(f"未找到依赖工具|Required tool not found: {tool}")

    def _generate_sample_mapping(self):
        """生成或验证样品映射|Generate or validate sample mapping"""
        if self.config.sample_map_file:
            # 使用现有的样品映射文件|Use existing sample mapping file
            self._load_sample_mapping()
        elif self.config.input:
            # 自动生成样品映射文件|Auto-generate sample mapping
            self._auto_generate_sample_mapping()
        else:
            raise ValueError("必须指定输入文件或样品映射文件|Must specify input file or sample mapping file")

    def _load_sample_mapping(self):
        """加载样品映射文件|Load sample mapping file"""
        self.logger.info(f"加载样品映射文件|Loading sample mapping file: {self.config.sample_map_file}")

        try:
            with open(self.config.sample_map_file, 'r') as f:
                for line_num, line in enumerate(f, 1):
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue

                    parts = line.split('\t')
                    if len(parts) >= 2:
                        file_path = parts[0].strip()
                        sample_name = parts[1].strip()

                        if os.path.exists(file_path):
                            self.sample_mapping[sample_name] = file_path
                        else:
                            self.logger.warning(f"文件不存在|File not found: {file_path}")
                    else:
                        self.logger.warning(f"第{line_num}行格式错误|Line {line_num} format error: {line}")

            if not self.sample_mapping:
                raise ValueError("样品映射文件为空或格式错误|Sample mapping file is empty or incorrectly formatted")

            self.logger.info(f"加载了{len(self.sample_mapping)}个样品|Loaded {len(self.sample_mapping)} samples")

        except Exception as e:
            raise RuntimeError(f"加载样品映射文件失败|Failed to load sample mapping file: {e}")

    def _auto_generate_sample_mapping(self):
        """自动生成样品映射|Auto-generate sample mapping"""
        self.logger.info(f"自动扫描输入文件|Auto-scanning input files: {self.config.input}")

        input_path = Path(self.config.input)

        if input_path.is_file():
            # 单个文件|Single file
            sample_name = self.config.sample_name or input_path.stem
            self.sample_mapping[sample_name] = str(input_path)
            self.config.auto_generated_map = True

        elif input_path.is_dir():
            # 目录,扫描匹配的文件|Directory, scan matching files
            pattern = self.config.input_suffix
            files = glob.glob(os.path.join(str(input_path), pattern))

            for file_path in files:
                file_name = os.path.basename(file_path)
                if self.config.auto_detect_samples:
                    # 自动检测:用正则提取样品名称|Auto-detect: extract sample name via regex
                    match = re.search(self.config.sample_name_pattern, file_name)
                    sample_name = match.group(1) if match else Path(file_path).stem
                else:
                    # 关闭自动检测:直接用文件名stem|Disabled: use filename stem directly
                    sample_name = Path(file_path).stem
                self.sample_mapping[sample_name] = file_path

            self.config.auto_generated_map = True

        else:
            raise FileNotFoundError(f"输入路径不存在|Input path not found: {self.config.input}")

        if not self.sample_mapping:
            raise ValueError(f"在{self.config.input}中未找到匹配的文件|No matching files found in {self.config.input}")

        self.logger.info(f"发现了{len(self.sample_mapping)}个样品|Found {len(self.sample_mapping)} samples")

    def run_analysis(self) -> bool:
        """
        运行BLAST分析流程|Run BLAST analysis pipeline

        Returns:
            bool: 分析是否成功|Whether analysis was successful
        """
        try:
            # 准备: 创建输出目录|Prep: Create output directory
            if not self._create_output_directory():
                return False

            # 记录软件版本信息到00_pipeline_info|Record software versions to 00_pipeline_info
            self._generate_software_versions()

            # 步骤1: 创建BLAST数据库|Step 1: Create BLAST database
            if not self._create_blast_database():
                return False

            # 步骤2: 运行BLAST比对|Step 2: Run BLAST alignment
            if not self._run_blast_alignment():
                return False

            # 步骤3: 处理结果|Step 3: Process results
            if not self._process_results():
                return False

            # 步骤4: 生成统计报告|Step 4: Generate statistics report
            if not self._generate_statistics_report():
                return False

            # 步骤5: 生成比对可视化|Step 5: Generate alignment visualization
            if self.config.alignment_output != 'none':
                if not self._generate_alignment_visualization():
                    return False

            return True

        except Exception as e:
            self.logger.error(f"BLAST分析失败|BLAST analysis failed: {e}", exc_info=True)
            return False

    def _create_output_directory(self) -> bool:
        """创建输出目录及分层子目录(§12.2)|Create output directory and layered subdirs"""
        try:
            for subdir in (self.config.output, self.config.pipeline_info_dir,
                           self.config.db_dir, self.config.blast_dir,
                           self.config.alignments_dir, self.config.logs_dir):
                os.makedirs(subdir, exist_ok=True)
            self.logger.info(f"输出目录已创建|Output directory created: {self.config.output}")
            return True
        except Exception as e:
            self.logger.error(f"创建输出目录失败|Failed to create output directory: {e}")
            return False

    def _generate_software_versions(self):
        """生成software_versions.yml到00_pipeline_info(§12.5)|Generate software_versions.yml"""
        import yaml

        blast_program = getattr(self.config, f'{self.config.blast_type}_path', self.config.blast_type)
        tools = {
            'makeblastdb': self.config.makeblastdb_path,
            self.config.blast_type: blast_program,
        }

        versions = {}
        for name, path in tools.items():
            try:
                cmd = build_conda_command(path, ['-version'])
                result = subprocess.run(cmd, capture_output=True, text=True, timeout=30)
                ver = (result.stdout or result.stderr or '').strip().split('\n')[0]
                versions[name] = {'version': ver or 'unknown', 'path': path}
            except Exception as e:
                self.logger.warning(f"获取{name}版本失败|Failed to get {name} version: {e}")
                versions[name] = {'version': 'unknown', 'path': path}

        info = {
            'pipeline': {'name': 'biopytools blast', 'version': '2.1.0'},
            'tools': versions,
            'parameters': {
                'blast_type': self.config.blast_type,
                'evalue': self.config.evalue,
                'word_size': self.config.word_size,
                'max_target_seqs': self.config.max_target_seqs,
                'target_db_type': self.config.target_db_type,
                'min_identity': self.config.min_identity,
                'min_coverage': self.config.min_coverage,
                'threads': self.config.threads,
            },
        }

        output_file = os.path.join(self.config.pipeline_info_dir, 'software_versions.yml')
        try:
            with open(output_file, 'w') as f:
                yaml.dump(info, f, default_flow_style=False, allow_unicode=True)
            self.logger.info(f"软件版本信息已保存|Software versions saved to: {output_file}")
        except Exception as e:
            self.logger.warning(f"保存软件版本信息失败|Failed to save software versions: {e}")

    def _create_blast_database(self) -> bool:
        """创建BLAST数据库|Create BLAST database"""
        try:
            self.log_step_start("创建BLAST数据库|Creating BLAST Database", 1)

            db_path = self.config.get_target_db_path()

            # 检查数据库是否已存在|Check if database already exists
            # makeblastdb生成.db.nhr(nucl)/.db.phr(prot),而非.db本身
            # |makeblastdb produces .db.nhr(nucl)/.db.phr(prot), not .db itself
            db_index_exists = (
                os.path.exists(f"{db_path}.nhr") or os.path.exists(f"{db_path}.phr")
            )
            if db_index_exists and not self.config.force:
                self.logger.info(f"数据库已存在，跳过创建|Database already exists, skipping creation: {db_path}")
                self.database_path = db_path
                self.log_step_end("创建BLAST数据库|Creating BLAST Database", True)
                return True

            # 构建数据库|Build database
            args = [
                '-in', self.config.reference,
                '-dbtype', self.config.target_db_type,
                '-out', db_path
            ]

            cmd = build_conda_command(self.config.makeblastdb_path, args)

            self.logger.info("执行|Executing: makeblastdb建库|makeblastdb database building")
            self.logger.info(f"命令|Command: {' '.join(cmd)}")

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                check=True
            )

            self.database_path = db_path
            self.logger.info(f"数据库创建成功|Database created successfully: {db_path}")

            self.log_step_end("创建BLAST数据库|Creating BLAST Database", True)
            return True

        except subprocess.CalledProcessError as e:
            self.logger.error(f"创建数据库失败|Database creation failed: {e}")
            if e.stdout:
                self.logger.error(f"stdout: {e.stdout}")
            if e.stderr:
                self.logger.error(f"stderr: {e.stderr}")
            return False
        except Exception as e:
            self.logger.error(f"创建数据库时发生错误|Error creating database: {e}")
            return False

    def _run_blast_alignment(self) -> bool:
        """运行BLAST比对|Run BLAST alignment"""
        try:
            self.log_step_start("运行BLAST比对|Running BLAST Alignment", 2)

            all_results = []
            total_samples = len(self.sample_mapping)

            for i, (sample_name, sample_file) in enumerate(self.sample_mapping.items(), 1):
                self.logger.info(f"处理样品 {i}/{total_samples}: {sample_name}|Processing sample {i}/{total_samples}: {sample_name}")

                result_file = self._run_single_blast(sample_name, sample_file)
                if result_file:
                    all_results.append((sample_name, result_file))
                else:
                    self.logger.warning(f"样品 {sample_name} 比对失败|Sample {sample_name} alignment failed")

            if not all_results:
                raise RuntimeError("所有样品比对都失败了|All sample alignments failed")

            # 合并结果|Merge results
            merged_file = self._merge_results(all_results)

            self.logger.info(f"BLAST比对完成|BLAST alignment completed. Results: {merged_file}")

            self.log_step_end("运行BLAST比对|Running BLAST Alignment", True)
            return True

        except Exception as e:
            self.logger.error(f"BLAST比对失败|BLAST alignment failed: {e}")
            self.log_step_end("运行BLAST比对|Running BLAST Alignment", False)
            return False

    def _run_single_blast(self, sample_name: str, sample_file: str) -> Optional[str]:
        """运行单个样品的BLAST比对|Run BLAST alignment for single sample"""
        try:
            output_file = os.path.join(self.config.blast_dir, f"{sample_name}_{self.config.blast_type}_results.tsv")

            # 断点续传:输出已存在且非空则跳过(§10.2)|Checkpoint: skip if output exists and non-empty
            if (not self.config.force and os.path.exists(output_file)
                    and os.path.getsize(output_file) > 0):
                self.logger.info(f"样品 {sample_name} 比对结果已存在,跳过|Sample {sample_name} result exists, skipping")
                return output_file

            # 获取正确的BLAST程序路径|Get correct BLAST program path
            blast_program = getattr(self.config, f'{self.config.blast_type}_path', self.config.blast_type)

            args = [
                '-query', sample_file,
                '-db', self.database_path,
                '-out', output_file,
                '-outfmt', '6 qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore slen qseq sseq',
                '-evalue', str(self.config.evalue),
                '-max_target_seqs', str(self.config.max_target_seqs),
                '-word_size', str(self.config.word_size),
                '-num_threads', str(self.config.threads)
            ]

            cmd = build_conda_command(blast_program, args)

            self.logger.info(f"执行|Executing: {self.config.blast_type}比对|{self.config.blast_type} alignment (sample: {sample_name})")
            self.logger.info(f"命令|Command: {' '.join(cmd)}")

            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                check=True
            )

            # 检查输出文件|Check output file
            if os.path.exists(output_file) and os.path.getsize(output_file) > 0:
                self.logger.info(f"样品 {sample_name} 比对完成|Sample {sample_name} alignment completed")
                return output_file
            else:
                self.logger.warning(f"样品 {sample_name} 比对输出为空|Sample {sample_name} alignment output is empty")
                return None

        except subprocess.CalledProcessError as e:
            self.logger.error(f"样品 {sample_name} BLAST比对失败|Sample {sample_name} BLAST alignment failed: {e}")
            if e.stderr:
                self.logger.error(f"stderr: {e.stderr}")
            return None
        except Exception as e:
            self.logger.error(f"样品 {sample_name} 比对时发生错误|Error during sample {sample_name} alignment: {e}")
            return None

    def _merge_results(self, result_files: List[Tuple[str, str]]) -> str:
        """合并所有结果文件并计算覆盖度|Merge all result files and calculate coverage"""
        try:
            merged_file = self.config.get_summary_output_path()

            with open(merged_file, 'w', encoding='utf-8') as out_f:
                # 写入中文表头|Write Chinese header (16列)
                out_f.write("样品名称\t查询序列ID\t目标序列ID\t序列相似度(%)\t比对长度\t错配数\tGap数\t")
                out_f.write("查询起始位置\t查询结束位置\t目标起始位置\t目标结束位置\tE-value\tBit_Score\t")
                out_f.write("目标序列长度\t目标序列覆盖度(%)\t查询序列\t目标序列\n")

                for sample_name, result_file in result_files:
                    # 直接使用传入的真实sample_name,避免从文件名反推(含下划线的样本名会被切错)
                    # |Use the real sample_name passed in; do not reverse-derive from filename (underscore names get split)

                    try:
                        with open(result_file, 'r', encoding='utf-8') as in_f:
                            for line in in_f:
                                line = line.strip()
                                if line:
                                    parts = line.split('\t')
                                    if len(parts) >= 14:
                                        try:
                                            # 计算覆盖度|Calculate coverage
                                            # BLAST output format (15 columns): qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore slen qseq sseq
                                            qseqid, sseqid, pident, length, mismatch, gapopen, qstart, qend, sstart, send, evalue, bitscore, slen, qseq, sseq = parts[:15]

                                            # 计算目标序列覆盖度|Calculate subject sequence coverage
                                            try:
                                                sstart_int = int(sstart)
                                                send_int = int(send)
                                                slen_int = int(slen)
                                                coverage = abs(send_int - sstart_int + 1) / slen_int * 100
                                                coverage = min(coverage, 100.0)
                                            except (ValueError, ZeroDivisionError):
                                                coverage = 0.0

                                            # 写入合并文件，添加覆盖度列|Write to merged file with coverage column
                                            out_line = '\t'.join([
                                                sample_name, qseqid, sseqid, pident, length, mismatch, gapopen,
                                                qstart, qend, sstart, send, evalue, bitscore, slen,
                                                f"{coverage:.2f}", qseq, sseq
                                            ])
                                            out_f.write(out_line + '\n')
                                        except (ValueError, IndexError) as e:
                                            self.logger.debug(f"跳过格式错误的行|Skipping malformed line: {e}")
                                            continue
                    except Exception as e:
                        self.logger.warning(f"读取结果文件失败|Failed to read result file: {result_file}, error: {e}")

            self.logger.info(f"合并结果文件|Merged results file: {merged_file}")
            return merged_file

        except Exception as e:
            raise RuntimeError(f"合并结果文件失败|Failed to merge result files: {e}")

    def _process_results(self) -> bool:
        """处理结果|Process results"""
        try:
            self.log_step_start("处理结果|Processing Results", 3)

            summary_file = self.config.get_summary_output_path()

            if not os.path.exists(summary_file):
                raise FileNotFoundError(f"汇总结果文件不存在|Summary result file not found: {summary_file}")

            # 统计结果|Generate statistics
            stats = self._generate_statistics(summary_file)
            self.log_statistics(stats)

            # 对结果排序|Sort results
            sorted_file = self._sort_results(summary_file)

            # 创建高质量结果文件|Create high quality results file
            self._create_high_quality_results(sorted_file)

            # 导出多sheet Excel(默认;软依赖pandas/openpyxl,缺失则跳过,TSV照常输出)|
            # Export multi-sheet Excel (default; soft-dep, skip if missing, TSV unaffected)
            self._export_excel(summary_file, sorted_file)

            self.log_step_end("处理结果|Processing Results", True)
            return True

        except Exception as e:
            self.logger.error(f"处理结果失败|Failed to process results: {e}")
            self.log_step_end("处理结果|Processing Results", False)
            return False

    def _export_excel(self, summary_file: str, sorted_file: str) -> bool:
        """导出多sheet Excel(软依赖pandas/openpyxl,缺失则跳过,不影响TSV输出)|Export multi-sheet Excel (soft-dep; skip if missing, TSV unaffected)"""
        # 软依赖:pandas/openpyxl 缺失则 warning + 跳过,流程继续(优雅降级,§graceful-degradation)|
        # Soft dep: warn + skip if missing, flow continues (graceful degradation)
        try:
            import pandas as pd
            import openpyxl  # noqa: F401  pandas 写 xlsx 的引擎|engine for pandas xlsx
        except ImportError:
            self.logger.warning("pandas/openpyxl 未安装,跳过 Excel 输出(TSV 照常输出)|pandas/openpyxl not installed, skip Excel (TSV still output)")
            return False

        import glob

        blast_dir = self.config.blast_dir
        xlsx_path = os.path.join(blast_dir, 'blast_results.xlsx')
        sheets = {}

        # raw_results:合并所有 {sample}_{blast_type}_results.tsv(blast outfmt 6 无表头,15列)|
        # merge all raw results (blast outfmt 6 has no header, 15 cols)
        raw_cols = ['qseqid', 'sseqid', 'pident', 'length', 'mismatch', 'gapopen',
                    'qstart', 'qend', 'sstart', 'send', 'evalue', 'bitscore',
                    'slen', 'qseq', 'sseq']
        pattern = os.path.join(blast_dir, f"*_{self.config.blast_type}_results.tsv")
        suffix = f"_{self.config.blast_type}_results.tsv"
        raw_dfs = []
        for rf in sorted(glob.glob(pattern)):
            try:
                df = pd.read_csv(rf, sep='\t', header=None, names=raw_cols, dtype=str)
                # 从文件名提取 sample,加 Sample 列区分多样本|extract sample, add Sample col
                base = os.path.basename(rf)
                sample = base[:-len(suffix)] if base.endswith(suffix) else base
                df.insert(0, 'Sample', sample)
                raw_dfs.append(df)
            except Exception as e:
                self.logger.warning(f"读取 raw results 跳过|Skip raw results {rf}: {e}")
        if raw_dfs:
            sheets['raw_results'] = pd.concat(raw_dfs, ignore_index=True)

        # summary / sorted:各对应一个 sheet | one sheet each
        for sheet_name, path in [('summary', summary_file), ('sorted', sorted_file)]:
            if path and os.path.exists(path):
                try:
                    sheets[sheet_name] = pd.read_csv(path, sep='\t', dtype=str)
                except Exception as e:
                    self.logger.warning(f"读取 {sheet_name} 跳过|Skip {sheet_name}: {e}")

        # high_quality 路径从 sorted 推导(与 _create_high_quality_results 一致)|
        # derive high_quality path from sorted (consistent with _create_high_quality_results)
        if sorted_file:
            if '_sorted.tsv' in sorted_file:
                hq_file = sorted_file.replace('_sorted.tsv', '_sorted_high_quality.tsv')
            else:
                hq_file = sorted_file.replace('.tsv', '_high_quality.tsv')
            if os.path.exists(hq_file):
                try:
                    sheets['high_quality'] = pd.read_csv(hq_file, sep='\t', dtype=str)
                except Exception as e:
                    self.logger.warning(f"读取 high_quality 跳过|Skip high_quality: {e}")

        if not sheets:
            self.logger.warning("无 TSV 可导出 Excel|No TSV available to export to Excel")
            return False

        # 样式常量(专业配色)|style constants (professional palette)
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        tab_colors = {  # sheet 标签色|sheet tab colors
            'raw_results': '00B050',    # 绿|green
            'summary': '1F4E78',        # 蓝|blue
            'sorted': 'ED7D31',         # 橙|orange
            'high_quality': '7030A0',   # 紫|purple
        }
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')  # 表头深蓝|header dark blue
        header_font = Font(color='FFFFFF', bold=True)  # 白字加粗|white bold
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='BFBFBF')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # 写 Excel + 美化;单 sheet 超行限则跳过 | write Excel + styling; skip sheet exceeding row limit
        written = 0
        try:
            with pd.ExcelWriter(xlsx_path) as writer:
                for sheet_name, df in sheets.items():
                    try:
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                        ws = writer.sheets[sheet_name]
                        # sheet 标签色 + 冻结首行 | tab color + freeze first row
                        ws.sheet_properties.tabColor = tab_colors.get(sheet_name)
                        ws.freeze_panes = 'A2'
                        # 列宽按内容自适应(采样前 100 行估算)|auto width (sample first 100 rows)
                        for col_idx, col in enumerate(df.columns, 1):
                            col_letter = ws.cell(row=1, column=col_idx).column_letter
                            sample_vals = [str(col)] + [str(v) for v in df[col].head(100)]
                            max_len = max((len(v) for v in sample_vals), default=8)
                            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 50)
                        # 表头样式(深蓝底白字加粗居中) + 全表细边框 | header style + thin border
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                            for cell in row:
                                cell.border = border
                                if cell.row == 1:
                                    cell.fill = header_fill
                                    cell.font = header_font
                                    cell.alignment = center
                        written += 1
                    except Exception as e:
                        self.logger.warning(f"sheet '{sheet_name}' 写入跳过(可能超 Excel 行限)|Skip sheet '{sheet_name}' (may exceed Excel row limit): {e}")
            self.logger.info(f"Excel 已输出|Excel exported: {xlsx_path} ({written} sheets)")
            return True
        except Exception as e:
            self.logger.error(f"Excel 输出失败|Excel export failed: {e}")
            return False

    def _generate_statistics(self, result_file: str) -> Dict:
        """生成统计信息|Generate statistics"""
        try:
            stats = {
                'total_alignments': 0,
                'samples_count': len(self.sample_mapping),
                'unique_queries': set(),
                'unique_subjects': set()
            }

            with open(result_file, 'r') as f:
                next(f)  # 跳过表头

                for line in f:
                    parts = line.strip().split('\t')
                    if len(parts) >= 13:  # 13列: Sample + 12列BLAST输出
                        stats['total_alignments'] += 1
                        stats['unique_queries'].add(parts[1])  # qseqid
                        stats['unique_subjects'].add(parts[2])  # sseqid

            stats['unique_queries'] = len(stats['unique_queries'])
            stats['unique_subjects'] = len(stats['unique_subjects'])

            return stats

        except Exception as e:
            self.logger.error(f"生成统计信息失败|Failed to generate statistics: {e}")
            return {'total_alignments': 0, 'samples_count': 0, 'unique_queries': 0, 'unique_subjects': 0}

    def _sort_results(self, summary_file: str) -> str:
        """对结果按覆盖度、相似度和E-value排序|Sort results by coverage, identity and E-value"""
        try:
            self.logger.info("按覆盖度、相似度和E-value排序结果|Sorting results by coverage, identity and E-value")

            sorted_file = summary_file.replace('.tsv', '_sorted.tsv')

            # 读取所有数据行
            header = None
            data_lines = []

            with open(summary_file, 'r', encoding='utf-8') as f:
                header = f.readline().strip()
                for line in f:
                    line = line.strip()
                    if line:
                        data_lines.append(line)

            if not data_lines:
                self.logger.warning("没有数据需要排序|No data to sort")
                return summary_file

            # 解析并排序
            def parse_line(line):
                parts = line.split('\t')
                # 新格式：16列 - Sample qseqid sseqid pident length mismatch gapopen qstart qend sstart send evalue bitscore slen coverage qseq sseq
                if len(parts) < 16:
                    return None

                try:
                    # 解析列索引
                    # 0: Sample, 1: qseqid, 2: sseqid, 3: pident, 4: length,
                    # 5: mismatch, 6: gapopen, 7: qstart, 8: qend, 9: sstart, 10: send,
                    # 11: evalue, 12: bitscore, 13: slen, 14: coverage, 15: qseq, 16: sseq
                    identity = float(parts[3])
                    coverage = float(parts[14])
                    evalue_str = parts[11]

                    # Parse e-value
                    try:
                        if evalue_str.startswith('e') or evalue_str.startswith('E'):
                            evalue = float(f"1{evalue_str}")
                        else:
                            evalue = float(evalue_str)
                    except Exception:
                        evalue = 1.0

                    return {
                        'line': line,
                        'coverage': coverage,
                        'identity': identity,
                        'evalue': evalue
                    }
                except (ValueError, IndexError):
                    return None

            # 解析所有行
            parsed_data = [parse_line(line) for line in data_lines]
            parsed_data = [d for d in parsed_data if d is not None]

            # 排序：覆盖度降序、相似度降序、E-value升序
            parsed_data.sort(key=lambda x: (-x['coverage'], -x['identity'], x['evalue']))

            # 写入排序后的文件
            with open(sorted_file, 'w', encoding='utf-8') as f:
                f.write(header + '\n')
                for item in parsed_data:
                    f.write(item['line'] + '\n')

            self.logger.info(f"排序完成|Sorting completed: {sorted_file}")
            self.logger.info(f"排序结果数|Sorted results: {len(parsed_data)}")

            return sorted_file

        except Exception as e:
            self.logger.error(f"排序失败|Sorting failed: {e}")
            return summary_file

    def _create_high_quality_results(self, sorted_file: str) -> bool:
        """从排序文件创建高质量比对结果文件|Create high quality alignment results from sorted file"""
        try:
            self.logger.info("从排序文件筛选高质量比对结果|Filtering high quality alignments from sorted file")

            # 从 sorted.tsv 生成 sorted_high_quality.tsv
            if '_sorted.tsv' in sorted_file:
                high_quality_file = sorted_file.replace('_sorted.tsv', '_sorted_high_quality.tsv')
            else:
                high_quality_file = sorted_file.replace('.tsv', '_high_quality.tsv')

            # 读取排序后的结果
            high_quality_alignments = []
            total_count = 0

            with open(sorted_file, 'r', encoding='utf-8') as f:
                # 读取表头
                header = f.readline().strip()

                for line in f:
                    line = line.strip()
                    if not line:
                        continue

                    total_count += 1
                    parts = line.split('\t')

                    # 新格式：16列
                    if len(parts) < 16:
                        continue

                    try:
                        # Sample, qseqid, sseqid, pident, length, mismatch, gapopen,
                        # qstart, qend, sstart, send, evalue, bitscore, slen, coverage, qseq, sseq
                        sample_name = parts[0]
                        query_id = parts[1]
                        subject_id = parts[2]
                        identity = float(parts[3])
                        length = int(parts[4])
                        evalue = parts[11]
                        coverage = float(parts[14])

                        # 解析e-value
                        try:
                            if evalue.startswith('e') or evalue.startswith('E'):
                                evalue_float = float(f"1{evalue}")
                            else:
                                evalue_float = float(evalue)
                        except Exception:
                            evalue_float = 1.0

                        # 筛选条件：E-value、相似度和覆盖度|Filter criteria: E-value, identity and coverage
                        if (evalue_float <= self.config.high_quality_evalue and
                            identity >= self.config.min_identity and
                            coverage >= self.config.min_coverage):
                            high_quality_alignments.append(line)

                    except (ValueError, IndexError) as e:
                        self.logger.debug(f"跳过格式错误的行|Skipping malformed line: {e}")
                        continue

            # 写入高质量结果文件
            if high_quality_alignments:
                with open(high_quality_file, 'w', encoding='utf-8') as f:
                    f.write(header + '\n')
                    for alignment in high_quality_alignments:
                        f.write(alignment + '\n')

                self.logger.info(f"高质量结果文件已创建|High quality results file created: {high_quality_file}")
                self.logger.info(f"高质量比对数量|High quality alignments: {len(high_quality_alignments)} / {total_count}")
            else:
                self.logger.warning(f"未找到符合条件的高质量比对结果|No high quality alignments found")
                self.logger.info(f"筛选条件|Filter criteria: E-value <= {self.config.high_quality_evalue}, "
                               f"Identity >= {self.config.min_identity}%, "
                               f"Coverage >= {self.config.min_coverage}%")

            return True

        except Exception as e:
            self.logger.error(f"创建高质量结果文件失败|Failed to create high quality results file: {e}")
            return False

    def _generate_statistics_report(self) -> bool:
        """生成统计报告|Generate statistics report"""
        try:
            from .statistics import StatisticsGenerator

            self.log_step_start("生成统计报告|Generating Statistics Report", 4)

            summary_file = self.config.get_summary_output_path()

            if not os.path.exists(summary_file) or os.path.getsize(summary_file) == 0:
                self.logger.warning("汇总结果文件为空，跳过统计报告生成|Summary file is empty, skipping statistics report")
                self.log_step_end("生成统计报告|Generating Statistics Report", True)
                return True

            stats_generator = StatisticsGenerator(self.config, self.logger)
            stats_file = stats_generator.generate_statistics_report(summary_file)

            if stats_file:
                self.log_step_end("生成统计报告|Generating Statistics Report", True)
                return True
            else:
                self.logger.warning("统计报告生成失败，但继续执行|Statistics report generation failed, but continuing")
                self.log_step_end("生成统计报告|Generating Statistics Report", True)
                return True

        except Exception as e:
            self.logger.error(f"生成统计报告失败|Failed to generate statistics report: {e}")
            self.log_step_end("生成统计报告|Generating Statistics Report", False)
            return False

    def _generate_alignment_visualization(self) -> bool:
        """生成比对可视化|Generate alignment visualization"""
        try:
            from .alignment_visualizer import AlignmentVisualizer

            self.log_step_start("生成比对可视化|Generating Alignment Visualization", 5)

            # 准备BLAST结果数据|Prepare BLAST result data
            blast_results = []
            for sample_name, sample_file in self.sample_mapping.items():
                result_file = os.path.join(self.config.blast_dir, f"{sample_name}_{self.config.blast_type}_results.tsv")
                if os.path.exists(result_file) and os.path.getsize(result_file) > 0:
                    blast_results.append((os.path.basename(sample_file), sample_name, result_file))

            if not blast_results:
                self.logger.warning("没有可用于可视化的比对结果|No alignments available for visualization")
                self.log_step_end("生成比对可视化|Generating Alignment Visualization", True)
                return True

            # 使用AlignmentVisualizer生成可视化|Use AlignmentVisualizer
            visualizer = AlignmentVisualizer(self.config, self.logger)
            output_files = visualizer.generate_visualizations(blast_results)

            if output_files:
                self.logger.info(f"比对可视化生成完成|Alignment visualization generated successfully")
                if 'html' in output_files:
                    self.logger.info(f"  HTML索引文件|HTML index file: {output_files['html']['index']}")
                if 'text' in output_files:
                    self.logger.info(f"  文本摘要文件|Text summary file: {output_files['text']['summary']}")
            else:
                self.logger.warning("比对可视化生成失败|Failed to generate alignment visualization")

            self.log_step_end("生成比对可视化|Generating Alignment Visualization", True)
            return True

        except Exception as e:
            self.logger.error(f"生成比对可视化失败|Failed to generate alignment visualization: {e}")
            self.log_step_end("生成比对可视化|Generating Alignment Visualization", False)
            return False

    def log_additional_summary(self):
        """记录额外的摘要信息|Log additional summary information"""
        self.logger.info(f"样品数量|Sample count: {len(self.sample_mapping)}")
        self.logger.info(f"BLAST类型|BLAST type: {self.config.blast_type}")
        self.logger.info(f"目标数据库|Target database: {self.config.reference}")
        self.logger.info(f"输出目录|Output directory: {self.config.output}")

        # 记录比对可视化信息|Log alignment visualization information
        if self.config.alignment_output != 'none':
            self.logger.info(f"比对可视化|Alignment visualization: {self.config.alignment_output}")
            if self.config.alignment_output in ['html', 'both']:
                html_path = self.config.get_alignment_output_path('html')
                self.logger.info(f"HTML可视化路径|HTML visualization path: {html_path}")
            if self.config.alignment_output in ['text', 'both']:
                text_path = self.config.get_alignment_output_path('text')
                self.logger.info(f"文本可视化路径|Text visualization path: {text_path}")

        if self.config.auto_generated_map:
            self.logger.info("使用自动生成的样品映射|Used auto-generated sample mapping")


def main():
    """命令行入口函数|Command line entry function"""
    import argparse

    parser = argparse.ArgumentParser(
        description="BLAST序列比对分析工具|BLAST Sequence Alignment Analysis Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )

    # 输入参数|Input arguments
    parser.add_argument('-i', '--input', default=None,
                       help='输入文件或目录路径|Input file or directory path')
    parser.add_argument('-s', '--sample-map-file', default=None,
                       help='样品映射文件(与-i二选一)|Sample mapping file (alternative to -i)')
    parser.add_argument('-r', '--reference', required=True,
                       help='目标基因序列文件|Target gene sequence file')

    # 输出参数|Output arguments
    parser.add_argument('-o', '--output-dir', default='./blast_output',
                       help='输出目录|Output directory')

    # BLAST参数|BLAST parameters
    parser.add_argument('--blast-type', default=None,
                       choices=['blastn', 'blastp', 'blastx', 'tblastn', 'tblastx'],
                       help='BLAST程序类型,默认自动检测|BLAST program type (auto-detect if not specified)')
    parser.add_argument('-e', '--evalue', type=float, default=1e-5,
                       help='E-value阈值|E-value threshold')
    parser.add_argument('--max-target-seqs', type=int, default=10,
                       help='最大目标序列数|Maximum target sequences')
    parser.add_argument('--word-size', type=int, default=None,
                       help='词大小,默认按blast-type设置|Word size (auto-set by blast-type)')
    parser.add_argument('-t', '--threads', type=int, default=12,
                       help='线程数|Number of threads')
    parser.add_argument('--input-suffix', default='*.fa',
                       help='输入文件后缀模式|Input file suffix pattern')
    parser.add_argument('--target-db-type', default=None,
                       choices=['nucl', 'prot'],
                       help='目标数据库类型,默认按blast-type设置|Target database type (auto-set by blast-type)')
    parser.add_argument('--min-identity', type=float, default=70.0,
                       help='最小序列相似度(%%)|Minimum sequence identity (%%)')
    parser.add_argument('--min-coverage', type=float, default=50.0,
                       help='最小覆盖度(%%)|Minimum coverage (%%)')
    parser.add_argument('--high-quality-evalue', type=float, default=1e-10,
                       help='高质量比对E-value阈值|High quality alignment E-value threshold')

    # 样品参数|Sample parameters
    parser.add_argument('--sample-name', default=None,
                       help='单文件输入时的样品名称|Sample name for single-file input')
    parser.add_argument('--no-auto-detect-samples', action='store_false',
                       dest='auto_detect_samples',
                       help='关闭自动检测样品名称|Disable auto sample name detection')
    parser.set_defaults(auto_detect_samples=True)
    parser.add_argument('--sample-name-pattern', default=r'([^/]+?)(?:\.fa|\.fasta|\.fna)?$',
                       help='样品名提取正则表达式|Sample name extraction regex')

    # 工具路径参数(默认None走get_tool_path取conda env完整路径)|Tool paths (None => get_tool_path resolves conda env)
    parser.add_argument('--makeblastdb-path', default=None,
                       help='makeblastdb程序路径|makeblastdb program path')
    parser.add_argument('--blastn-path', default=None,
                       help='blastn程序路径|blastn program path')
    parser.add_argument('--blastp-path', default=None,
                       help='blastp程序路径|blastp program path')
    parser.add_argument('--blastx-path', default=None,
                       help='blastx程序路径|blastx program path')
    parser.add_argument('--tblastn-path', default=None,
                       help='tblastn程序路径|tblastn program path')
    parser.add_argument('--tblastx-path', default=None,
                       help='tblastx程序路径|tblastx program path')

    # 比对可视化参数|Alignment visualization parameters
    parser.add_argument('--alignment-output', default='both',
                       choices=['none', 'text', 'html', 'both'],
                       help='比对可视化输出格式|Alignment visualization output format')
    parser.add_argument('--alignment-width', type=int, default=80,
                       help='比对每行显示的字符数|Characters per line in alignment display')
    parser.add_argument('--alignment-min-identity', type=float, default=0.0,
                       help='比对可视化最小相似度过滤|Minimum identity for alignment visualization')
    parser.add_argument('--alignment-min-coverage', type=float, default=0.0,
                       help='比对可视化最小覆盖度过滤|Minimum coverage for alignment visualization')
    parser.add_argument('--alignment-max-per-sample', type=int, default=100,
                       help='每个样品最多显示的比对数|Maximum alignments to display per sample')
    parser.add_argument('--html-theme', default='modern',
                       choices=['modern', 'classic', 'dark'],
                       help='HTML主题样式|HTML theme style')

    # 日志与执行控制|Logging and execution control
    parser.add_argument('-v', '--verbose', action='count', default=0,
                       help='详细输出(-vv更详细)|Verbose output (-vv for more)')
    parser.add_argument('--quiet', action='store_true',
                       help='静默模式|Quiet mode')
    parser.add_argument('--log-level', default='INFO',
                       choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                       help='日志级别|Log level')
    parser.add_argument('--log-file', default=None,
                       help='日志文件路径|Log file path')
    parser.add_argument('-f', '--force', action='store_true',
                       help='强制覆盖已存在文件|Force overwrite existing files')
    parser.add_argument('--dry-run', action='store_true',
                       help='模拟运行不执行|Dry run without execution')

    args = parser.parse_args()

    # 日志级别推导:显式--log-level优先,否则按-v推导|Derive log level: explicit --log-level wins, else from -v
    if '--log-level' not in sys.argv:
        if args.verbose >= 2:
            args.log_level = 'DEBUG'
        elif args.verbose == 1:
            args.log_level = 'INFO'
        else:
            args.log_level = 'WARNING'

    # 创建配置|Create configuration
    config = BLASTConfig(
        input=args.input,
        reference=args.reference,
        output_dir=args.output_dir,
        sample_name=args.sample_name,
        sample_map_file=args.sample_map_file,
        threads=args.threads,
        evalue=args.evalue,
        max_target_seqs=args.max_target_seqs,
        word_size=args.word_size,
        input_suffix=args.input_suffix,
        target_db_type=args.target_db_type,
        blast_type=args.blast_type,
        min_identity=args.min_identity,
        min_coverage=args.min_coverage,
        high_quality_evalue=args.high_quality_evalue,
        auto_detect_samples=args.auto_detect_samples,
        sample_name_pattern=args.sample_name_pattern,
        makeblastdb_path=args.makeblastdb_path,
        blastn_path=args.blastn_path,
        blastp_path=args.blastp_path,
        blastx_path=args.blastx_path,
        tblastn_path=args.tblastn_path,
        tblastx_path=args.tblastx_path,
        alignment_output=args.alignment_output,
        alignment_width=args.alignment_width,
        alignment_min_identity=args.alignment_min_identity,
        alignment_min_coverage=args.alignment_min_coverage,
        alignment_max_per_sample=args.alignment_max_per_sample,
        html_theme=args.html_theme,
        log_level=args.log_level,
        log_file=args.log_file,
        force=args.force,
        dry_run=args.dry_run,
    )

    # 创建分析器并运行|Create analyzer and run
    analyzer = BLASTAnalyzer(config)
    success = analyzer.run_pipeline()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()