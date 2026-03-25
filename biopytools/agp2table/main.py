"""
AGP转表格主程序模块|AGP to Table Main Module
"""

import sys
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from .config import AGPConfig
from .utils import AGPLogger


class AGPConverter:
    """AGP转表格转换器|AGP to Table Converter"""

    def __init__(self, **kwargs):
        # 初始化配置|Initialize configuration
        self.config = AGPConfig(**kwargs)
        self.config.validate()

        # 初始化日志|Initialize logging
        log_file = Path(self.config.output_file).parent / "agp2table.log"
        self.logger_manager = AGPLogger(log_file)
        self.logger = self.logger_manager.get_logger()

    def parse_agp(self) -> Tuple[List[Dict], Dict[str, List[Dict]], Dict[str, int]]:
        """
        解析AGP文件|Parse AGP file

        Returns:
            Tuple[List[Dict], Dict[str, List[Dict]], Dict[str, int]]: (所有记录, 按scaffold分组的记录, 每条scaffold的gap数量)
        """
        self.logger.info(f"解析AGP文件|Parsing AGP file: {self.config.agp_file}")

        records = []
        scaffold_groups = defaultdict(list)
        scaffold_gap_counts = defaultdict(int)

        with open(self.config.agp_file, 'r') as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                if not line or line.startswith('#'):
                    continue

                fields = line.split('\t')
                if len(fields) < 9:
                    self.logger.warning(f"行{line_num}格式不正确，跳过|Line {line_num} malformed, skipping")
                    continue

                # 解析AGP字段|Parse AGP fields
                record = {
                    'object': fields[0],          # scaffold/chromosome名称
                    'object_beg': int(fields[1]),  # 起始位置
                    'object_end': int(fields[2]),   # 结束位置
                    'part_number': fields[3],       # 部分编号
                    'component_type': fields[4],    # 组件类型 (W/U/N/O等)
                    'component_id': fields[5] if len(fields) > 5 else '',  # 组件ID
                    'component_beg': int(fields[6]) if len(fields) > 6 and fields[6].replace('.', '').isdigit() else 0,  # 组件起始
                    'component_end': int(fields[7]) if len(fields) > 7 and fields[7].replace('.', '').isdigit() else 0,   # 组件结束
                    'orientation': fields[8] if len(fields) > 8 else '',  # 方向 (+/-/?)
                    'line_number': line_num
                }

                # 计算长度|Calculate length
                record['length'] = record['object_end'] - record['object_beg'] + 1

                records.append(record)
                scaffold_groups[record['object']].append(record)

                # 统计gap数量|Count gaps
                if record['component_type'] == 'U':
                    scaffold_gap_counts[record['object']] += 1

        self.logger.info(f"解析完成|Parsing completed: {len(records)} 条记录|records, {len(scaffold_groups)} 个scaffold|scaffolds")
        return records, scaffold_groups, scaffold_gap_counts

    def calculate_statistics(self, records: List[Dict], scaffold_groups: Dict[str, List[Dict]]) -> Dict:
        """
        计算统计信息|Calculate statistics

        Returns:
            Dict: 统计信息字典
        """
        self.logger.info("计算统计信息|Calculating statistics")

        stats = {
            'total_records': len(records),
            'total_scaffolds': len(scaffold_groups),
            'total_length': 0,
            'components_by_type': defaultdict(int),
            'scaffold_lengths': {},
            'gap_count': 0,
            'gap_total_length': 0,
            'contig_count': 0
        }

        # 统计组件类型|Count component types
        for record in records:
            stats['components_by_type'][record['component_type']] += 1

        # 统计每个scaffold的长度|Calculate scaffold lengths
        for scaffold_name, scaffold_records in scaffold_groups.items():
            scaffold_length = max(r['object_end'] for r in scaffold_records)
            stats['scaffold_lengths'][scaffold_name] = scaffold_length
            stats['total_length'] += scaffold_length

        # 统计gap和contig|Count gaps and contigs
        for record in records:
            if record['component_type'] == 'U':  # gap
                stats['gap_count'] += 1
                stats['gap_total_length'] += record['length']
            elif record['component_type'] == 'W':  # contig
                stats['contig_count'] += 1

        return stats

    def format_table_line(self, record: Dict, gap_count: int) -> str:
        """
        格式化表格行|Format table line

        Args:
            record: AGP记录字典
            gap_count: 该scaffold的gap数量

        Returns:
            str: 格式化的表格行
        """
        if self.config.format == 'csv':
            separator = ','
        else:  # txt or tsv
            separator = '\t'

        # 类型描述|Type description
        type_descriptions = {
            'W': 'Contig',
            'U': 'Gap',
            'N': 'Gap',
            'O': 'Other'
        }
        component_type_desc = type_descriptions.get(record['component_type'], record['component_type'])

        line = separator.join([
            record['object'],
            str(record['object_beg']),
            str(record['object_end']),
            str(record['length']),
            str(gap_count),
            record['part_number'],
            component_type_desc,
            record.get('component_id', ''),
            str(record.get('component_beg', '')),
            str(record.get('component_end', '')),
            record.get('orientation', '')
        ])

        return line

    def write_excel(self, records: List[Dict], scaffold_groups: Dict[str, List[Dict]],
                    scaffold_gap_counts: Dict[str, int], stats: Dict) -> bool:
        """
        写入Excel格式文件|Write Excel format file

        Args:
            records: 所有记录
            scaffold_groups: 按scaffold分组的记录
            scaffold_gap_counts: 每条scaffold的gap数量
            stats: 统计信息

        Returns:
            bool: 是否成功
        """
        if not HAS_OPENPYXL:
            self.logger.error("openpyxl未安装，无法写入Excel格式|openpyxl not installed, cannot write Excel format")
            self.logger.error("请安装|Please install: pip install openpyxl")
            return False

        try:
            self.logger.info(f"创建Excel工作簿|Creating Excel workbook: {self.config.output_file}")

            # 创建工作簿|Create workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "AGP Data"

            # 定义表头|Define headers
            headers = ['Scaffold', 'Start', 'End', 'Length', 'Gap_Count',
                      'Part_Number', 'Type', 'Component_ID',
                      'Component_Start', 'Component_End', 'Orientation']

            # 写入表头|Write headers
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 冻结首行|Freeze first row
            ws.freeze_panes = 'A2'

            # 写入数据|Write data
            row_num = 2
            if self.config.group_by_scaffold:
                sorted_scaffolds = sorted(scaffold_groups.keys())
                for scaffold in sorted_scaffolds:
                    gap_count = scaffold_gap_counts.get(scaffold, 0)
                    for record in scaffold_groups[scaffold]:
                        self._write_excel_row(ws, row_num, record, gap_count)
                        row_num += 1
            else:
                for record in records:
                    gap_count = scaffold_gap_counts.get(record['object'], 0)
                    self._write_excel_row(ws, row_num, record, gap_count)
                    row_num += 1

            # 自动调整列宽|Auto-adjust column widths
            column_widths = {
                'A': 20,  # Scaffold
                'B': 12,  # Start
                'C': 12,  # End
                'D': 12,  # Length
                'E': 12,  # Gap_Count
                'F': 14,  # Part_Number
                'G': 10,  # Type
                'H': 20,  # Component_ID
                'I': 16,  # Component_Start
                'J': 16,  # Component_End
                'K': 14   # Orientation
            }

            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width

            # 如果有统计信息，添加到第二个sheet|Add statistics to second sheet if available
            if self.config.add_statistics and stats:
                self._write_statistics_sheet(wb, stats)

            # 保存工作簿|Save workbook
            wb.save(self.config.output_file)
            self.logger.info(f"Excel文件已保存|Excel file saved: {self.config.output_file}")

            return True

        except Exception as e:
            self.logger.error(f"写入Excel文件失败|Failed to write Excel file: {str(e)}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return False

    def _write_excel_row(self, ws, row_num: int, record: Dict, gap_count: int):
        """
        写入Excel行数据|Write Excel row data

        Args:
            ws: 工作表对象
            row_num: 行号
            record: AGP记录
            gap_count: gap数量
        """
        # 类型描述|Type description
        type_descriptions = {
            'W': 'Contig',
            'U': 'Gap',
            'N': 'Gap',
            'O': 'Other'
        }
        component_type_desc = type_descriptions.get(record['component_type'], record['component_type'])

        # 写入数据|Write data
        ws.cell(row=row_num, column=1, value=record['object'])
        ws.cell(row=row_num, column=2, value=record['object_beg'])
        ws.cell(row=row_num, column=3, value=record['object_end'])
        ws.cell(row=row_num, column=4, value=record['length'])
        ws.cell(row=row_num, column=5, value=gap_count)
        ws.cell(row=row_num, column=6, value=record['part_number'])
        ws.cell(row=row_num, column=7, value=component_type_desc)
        ws.cell(row=row_num, column=8, value=record.get('component_id', ''))
        ws.cell(row=row_num, column=9, value=record.get('component_beg', ''))
        ws.cell(row=row_num, column=10, value=record.get('component_end', ''))
        ws.cell(row=row_num, column=11, value=record.get('orientation', ''))

        # 根据类型设置不同颜色|Set different colors based on type
        if record['component_type'] == 'W':  # Contig
            fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        elif record['component_type'] == 'U':  # Gap
            fill = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
        else:
            fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        for col in range(1, 12):
            ws.cell(row=row_num, column=col).fill = fill
            ws.cell(row=row_num, column=col).alignment = Alignment(vertical='center')

    def _write_statistics_sheet(self, wb, stats: Dict):
        """
        写入统计信息到单独的sheet|Write statistics to separate sheet

        Args:
            wb: 工作簿对象
            stats: 统计信息字典
        """
        ws_stats = wb.create_sheet(title="Statistics")

        # 设置标题|Set title
        ws_stats.cell(row=1, column=1, value="统计信息|Statistics")
        ws_stats.cell(row=1, column=1).font = Font(bold=True, size=14)
        ws_stats.merge_cells('A1:B1')

        row = 3

        # 基本信息|Basic information
        info_data = [
            ("总记录数|Total records", stats['total_records']),
            ("总scaffold数|Total scaffolds", stats['total_scaffolds']),
            ("总长度|Total length", f"{stats['total_length']:,} bp"),
            ("", ""),
            ("Contig数量|Contig count", stats['contig_count']),
            ("Gap数量|Gap count", stats['gap_count']),
            ("Gap总长度|Gap total length", f"{stats['gap_total_length']:,} bp"),
        ]

        for label, value in info_data:
            if label:  # 跳过空行
                ws_stats.cell(row=row, column=1, value=label)
                ws_stats.cell(row=row, column=2, value=value)
                ws_stats.cell(row=row, column=1).font = Font(bold=True)
            row += 1

        row += 1

        # 组件类型统计|Component type statistics
        ws_stats.cell(row=row, column=1, value="组件类型统计|Component type statistics")
        ws_stats.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

        for comp_type, count in sorted(stats['components_by_type'].items()):
            ws_stats.cell(row=row, column=1, value=f"  {comp_type}")
            ws_stats.cell(row=row, column=2, value=count)
            row += 1

        row += 2

        # 前20个最长的scaffold|Top 20 longest scaffolds
        ws_stats.cell(row=row, column=1, value="前20个最长的scaffold|Top 20 longest scaffolds")
        ws_stats.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 2

        ws_stats.cell(row=row, column=1, value="排名|Rank")
        ws_stats.cell(row=row, column=2, value="Scaffold")
        ws_stats.cell(row=row, column=3, value="长度|Length (bp)")
        ws_stats.cell(row=row, column=1).font = Font(bold=True)
        ws_stats.cell(row=row, column=2).font = Font(bold=True)
        ws_stats.cell(row=row, column=3).font = Font(bold=True)
        row += 1

        sorted_scaffolds = sorted(stats['scaffold_lengths'].items(), key=lambda x: x[1], reverse=True)[:20]
        for i, (scaffold, length) in enumerate(sorted_scaffolds, 1):
            ws_stats.cell(row=row, column=1, value=i)
            ws_stats.cell(row=row, column=2, value=scaffold)
            ws_stats.cell(row=row, column=3, value=f"{length:,}")
            row += 1

        # 设置列宽|Set column widths
        ws_stats.column_dimensions['A'].width = 35
        ws_stats.column_dimensions['B'].width = 20
        ws_stats.column_dimensions['C'].width = 20

    def write_statistics(self, stats: Dict, output_file):
        """写入统计信息|Write statistics"""
        self.logger.info("写入统计信息|Writing statistics")

        with open(output_file, 'a', encoding='utf-8') as f:
            f.write("\n" + "=" * 100 + "\n")
            f.write("统计信息|Statistics\n")
            f.write("=" * 100 + "\n\n")

            f.write(f"总记录数|Total records: {stats['total_records']}\n")
            f.write(f"总scaffold数|Total scaffolds: {stats['total_scaffolds']}\n")
            f.write(f"总长度|Total length: {stats['total_length']:,} bp\n\n")

            f.write(f"Contig数量|Contig count: {stats['contig_count']}\n")
            f.write(f"Gap数量|Gap count: {stats['gap_count']}\n")
            f.write(f"Gap总长度|Gap total length: {stats['gap_total_length']:,} bp\n\n")

            f.write("组件类型统计|Component type statistics:\n")
            for comp_type, count in sorted(stats['components_by_type'].items()):
                f.write(f"  {comp_type}: {count}\n")

            f.write("\n前20个最长的scaffold|Top 20 longest scaffolds:\n")
            sorted_scaffolds = sorted(stats['scaffold_lengths'].items(), key=lambda x: x[1], reverse=True)[:20]
            for i, (scaffold, length) in enumerate(sorted_scaffolds, 1):
                f.write(f"  {i:2d}. {scaffold}: {length:,} bp\n")

    def convert(self) -> bool:
        """
        执行转换|Execute conversion

        Returns:
            bool: 是否成功|Success status
        """
        try:
            self.logger.info("开始AGP转表格|Starting AGP to table conversion")

            # 解析AGP文件|Parse AGP file
            records, scaffold_groups, scaffold_gap_counts = self.parse_agp()

            if not records:
                self.logger.error("AGP文件为空或格式不正确|AGP file is empty or malformed")
                return False

            # 计算统计信息|Calculate statistics
            stats = self.calculate_statistics(records, scaffold_groups) if self.config.add_statistics else {}

            # 如果是Excel格式，使用专门的Excel写入方法|If Excel format, use dedicated Excel writer
            if self.config.format == 'xlsx':
                return self.write_excel(records, scaffold_groups, scaffold_gap_counts, stats)

            # 写入输出文件|Write output file
            self.logger.info(f"写入输出文件|Writing output file: {self.config.output_file}")

            with open(self.config.output_file, 'w', encoding='utf-8') as f:
                # 写入表头|Write headers
                if self.config.add_headers:
                    if self.config.format == 'csv':
                        separator = ','
                    else:
                        separator = '\t'
                    header = separator.join([
                        'Scaffold', 'Start', 'End', 'Length', 'Gap_Count',
                        'Part_Number', 'Type', 'Component_ID',
                        'Component_Start', 'Component_End', 'Orientation'
                    ])
                    f.write(header + '\n')

                # 如果按scaffold分组，先排序|If group by scaffold, sort first
                if self.config.group_by_scaffold:
                    sorted_scaffolds = sorted(scaffold_groups.keys())
                    for scaffold in sorted_scaffolds:
                        gap_count = scaffold_gap_counts.get(scaffold, 0)
                        for record in scaffold_groups[scaffold]:
                            f.write(self.format_table_line(record, gap_count) + '\n')
                else:
                    for record in records:
                        gap_count = scaffold_gap_counts.get(record['object'], 0)
                        f.write(self.format_table_line(record, gap_count) + '\n')

                # 写入统计信息|Write statistics
                if self.config.add_statistics and stats:
                    f.write("\n" + "=" * 100 + "\n")
                    f.write("统计信息|Statistics\n")
                    f.write("=" * 100 + "\n\n")

                    f.write(f"总记录数|Total records: {stats['total_records']}\n")
                    f.write(f"总scaffold数|Total scaffolds: {stats['total_scaffolds']}\n")
                    f.write(f"总长度|Total length: {stats['total_length']:,} bp\n\n")

                    f.write(f"Contig数量|Contig count: {stats['contig_count']}\n")
                    f.write(f"Gap数量|Gap count: {stats['gap_count']}\n")
                    f.write(f"Gap总长度|Gap total length: {stats['gap_total_length']:,} bp\n\n")

                    f.write("组件类型统计|Component type statistics:\n")
                    for comp_type, count in sorted(stats['components_by_type'].items()):
                        f.write(f"  {comp_type}: {count}\n")

                    f.write("\n前20个最长的scaffold|Top 20 longest scaffolds:\n")
                    sorted_scaffolds = sorted(stats['scaffold_lengths'].items(), key=lambda x: x[1], reverse=True)[:20]
                    for i, (scaffold, length) in enumerate(sorted_scaffolds, 1):
                        f.write(f"  {i:2d}. {scaffold}: {length:,} bp\n")

            self.logger.info("转换完成|Conversion completed successfully")
            return True

        except Exception as e:
            self.logger.error(f"转换失败|Conversion failed: {str(e)}")
            import traceback
            self.logger.debug(traceback.format_exc())
            return False


def main():
    """主函数|Main function"""
    import argparse

    parser = argparse.ArgumentParser(
        description='AGP转表格工具|AGP to Table Converter',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )

    # 必需参数|Required arguments
    parser.add_argument('-i', '--input', required=True,
                       help='AGP文件路径|AGP file path')
    parser.add_argument('-o', '--output', required=True,
                       help='输出表格文件路径|Output table file path')

    # 可选参数|Optional arguments
    parser.add_argument('-f', '--format',
                       default='txt',
                       choices=['txt', 'tsv', 'csv', 'xlsx'],
                       help='输出格式|Output format')
    parser.add_argument('--statistics', action='store_true',
                       help='添加统计信息|Add statistics')
    parser.add_argument('--no-headers', action='store_true',
                       help='不添加表头|Do not add headers')
    parser.add_argument('--no-grouping', action='store_true',
                       help='不按scaffold分组|Do not group by scaffold')

    args = parser.parse_args()

    # 创建转换器并运行|Create converter and run
    converter = AGPConverter(
        agp_file=args.input,
        output_file=args.output,
        format=args.format,
        add_statistics=args.statistics,
        add_headers=not args.no_headers,
        group_by_scaffold=not args.no_grouping
    )

    success = converter.convert()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
