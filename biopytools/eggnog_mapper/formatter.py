"""
eggnog-mapper 注释结果重排版|eggnog-mapper annotations reformatter.
原生 .emapper.annotations -> 中文 TSV + Excel(双语表头).
"""

import os
from typing import List, Tuple

# 双语表头映射(未命中的列原样保留)|Bilingual header map
COLUMN_ZH = {
    "query": "查询ID|Query",
    "seed_ortholog": "种子直系同源|Seed ortholog",
    "evalue": "E值|E-value",
    "score": "比对得分|Score",
    "eggNOG_OGs": "eggNOG直系群|eggNOG OGs",
    "max_annot_lvl": "最大注释层级|Max annot level",
    "COG_category": "COG功能分类|COG category",
    "Description": "功能描述|Description",
    "Preferred_name": "推荐基因名|Preferred name",
    "GOs": "GO词条|GO terms",
    "EC_number": "EC号|EC number",
    "KEGG_ko": "KEGG_KO",
    "KEGG_Pathway": "KEGG通路|KEGG Pathway",
    "KEGG_Module": "KEGG模块|KEGG Module",
    "KEGG_Reaction": "KEGG反应|KEGG Reaction",
    "BRITE": "BRITE",
    "KEGG_TC": "KEGG_TC",
    "CAZy": "CAZy",
    "BiGG_Reaction": "BiGG反应|BiGG Reaction",
    "PFAMs": "Pfam",
}


def parse_annotations(path: str) -> Tuple[List[str], List[List[str]]]:
    """
    解析 .emapper.annotations|Parse annotations file.

    Returns:
        (headers, rows): headers 来自 #query 行;rows 为数据行(按列拆分)。
    """
    headers: List[str] = []
    rows: List[List[str]] = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue
            if line.startswith("#query"):
                headers = line.lstrip("#").split("\t")
            elif line.startswith("#"):
                continue
            else:
                rows.append(line.split("\t"))
    return headers, rows


class AnnotationsReformatter:
    """注释重排版器|Annotations reformatter (TSV + Excel)."""

    def __init__(self, annotations_file: str, output_dir: str, logger=None):
        self.annotations_file = annotations_file
        self.output_dir = output_dir
        self.logger = logger

        base = os.path.basename(annotations_file)
        suffix = ".emapper.annotations"
        if base.endswith(suffix):
            stem = base[: -len(suffix)]
        else:
            stem = os.path.splitext(base)[0]
        self.tsv_out = os.path.join(output_dir, f"{stem}.emapper.annotations.cn.tsv")
        self.xlsx_out = os.path.join(output_dir, f"{stem}.emapper.annotations.xlsx")

    def format(self) -> bool:
        """执行重排版 -> TSV + Excel|Run reformat."""
        headers, rows = parse_annotations(self.annotations_file)
        if not headers:
            if self.logger:
                self.logger.warning("未找到注释表头|No #query header in annotations")
            return False

        zh_headers = [COLUMN_ZH.get(h, h) for h in headers]
        self._write_tsv(zh_headers, rows)
        self._write_xlsx(zh_headers, rows)

        annotated = sum(1 for r in rows if any(c and c != "-" for c in r))
        if self.logger:
            self.logger.info(
                f"注释条数|Total queries: {len(rows)}, "
                f"有注释|with annotation: {annotated}"
            )
        return True

    def _write_tsv(self, headers: List[str], rows: List[List[str]]):
        """写中文 TSV|Write bilingual TSV."""
        with open(self.tsv_out, "w", encoding="utf-8") as f:
            f.write("\t".join(headers) + "\n")
            for r in rows:
                f.write("\t".join(r) + "\n")
        if self.logger:
            self.logger.info(f"中文TSV已保存|TSV saved: {self.tsv_out}")

    def _write_xlsx(self, headers: List[str], rows: List[List[str]]):
        """写 Excel(缺 openpyxl 则 WARNING 跳过)|Write Excel or skip gracefully."""
        try:
            import openpyxl
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            if self.logger:
                self.logger.warning(
                    "未安装openpyxl,跳过Excel|openpyxl missing, skip Excel"
                )
            return False

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "emapper注释|Annotations"

        bold = Font(bold=True)
        fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
        center = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = bold
            cell.fill = fill
            cell.alignment = center
        for ri, r in enumerate(rows, start=2):
            for ci, v in enumerate(r, start=1):
                ws.cell(row=ri, column=ci, value=v)

        ws.freeze_panes = "A2"
        wb.save(self.xlsx_out)
        if self.logger:
            self.logger.info(f"Excel已保存|Excel saved: {self.xlsx_out}")
        return True
