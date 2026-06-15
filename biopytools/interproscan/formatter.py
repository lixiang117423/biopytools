"""
InterProScan结果格式化输出模块|InterProScan Result Formatter Module
生成Excel和CSV格式的整理报告|Generate formatted Excel and CSV reports
"""

import logging
from pathlib import Path
from typing import Optional
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class InterProScanFormatter:
    """InterProScan结果格式化器|InterProScan Result Formatter"""

    def __init__(self, logger: logging.Logger):
        self.logger = logger

    def generate_report(
        self,
        parser,
        output_prefix: str,
        format_type: str = 'excel',
        include_summary: bool = True
    ) -> bool:
        """
        生成整理报告|Generate formatted report

        Args:
            parser: InterProScanParser对象|InterProScanParser object
            output_prefix: 输出文件前缀|Output file prefix
            format_type: 报告格式 ('excel', 'csv', 'both')|Report format
            include_summary: 是否包含汇总表|Whether to include summary sheet
        """
        try:
            if format_type.lower() == 'excel' or format_type.lower() == 'both':
                if not EXCEL_AVAILABLE:
                    self.logger.warning("openpyxl未安装，跳过Excel生成|openpyxl not installed, skip Excel generation")
                else:
                    self._generate_excel(parser, output_prefix, include_summary)

            if format_type.lower() == 'csv' or format_type.lower() == 'both':
                self._generate_csv(parser, output_prefix, include_summary)

            return True

        except Exception as e:
            self.logger.error(f"报告生成失败|Report generation failed: {str(e)}")
            return False

    def _generate_excel(self, parser, output_prefix: str, include_summary: bool):
        """生成Excel报告|Generate Excel report"""
        self.logger.info("生成Excel报告|Generating Excel report")

        output_file = f"{output_prefix}_formatted_report.xlsx"

        # 创建Excel工作簿|Create Excel workbook
        wb = Workbook()
        wb.remove(wb.active)  # 移除默认sheet|Remove default sheet

        sheet_index = 0

        # 1. 详细结果Sheet|Detailed Results Sheet
        if parser.matches:
            ws_matches = wb.create_sheet("Detailed Results", sheet_index)
            df_matches = parser.get_matches_dataframe()
            self._write_dataframe_to_sheet(ws_matches, df_matches, "详细注释结果|Detailed Annotation Results")
            sheet_index += 1

        # 2. GO注释Sheet（展开格式，每个GO一行）|GO Annotation Sheet (expanded, one GO per row)
        if parser.parse_goterms:
            df_go_expanded = parser.get_expanded_go_dataframe()
            if df_go_expanded is not None and not df_go_expanded.empty:
                ws_go = wb.create_sheet("GO Annotations", sheet_index)
                self._write_dataframe_to_sheet(ws_go, df_go_expanded, "GO注释(展开格式)|GO Annotations (Expanded Format)")
                sheet_index += 1

        # 3. Pathway注释Sheet（展开格式，每个Pathway一行）|Pathway Annotation Sheet (expanded, one Pathway per row)
        if parser.parse_pathways:
            df_pathway_expanded = parser.get_expanded_pathway_dataframe()
            if df_pathway_expanded is not None and not df_pathway_expanded.empty:
                ws_pathway = wb.create_sheet("Pathway Annotations", sheet_index)
                self._write_dataframe_to_sheet(ws_pathway, df_pathway_expanded, "Pathway注释(展开格式)|Pathway Annotations (Expanded Format)")
                sheet_index += 1

        # 4. 汇总统计Sheet|Summary Statistics Sheet
        if include_summary and parser.summaries:
            ws_summary = wb.create_sheet("Protein Summary", sheet_index)
            df_summary = parser.get_summaries_dataframe()
            # 排序：按匹配数降序|Sort by match count descending
            df_summary = df_summary.sort_values('Total Matches', ascending=False)
            self._write_dataframe_to_sheet(ws_summary, df_summary, "蛋白质注释汇总|Protein Annotation Summary")
            sheet_index += 1

        # 5. 数据库统计Sheet|Database Statistics Sheet
        if include_summary:
            ws_db = wb.create_sheet("Database Statistics", sheet_index)
            db_stats = parser.get_database_statistics()
            df_db = pd.DataFrame(list(db_stats.items()), columns=['Database', 'Match Count'])
            self._write_dataframe_to_sheet(ws_db, df_db, "数据库注释统计|Database Annotation Statistics")
            sheet_index += 1

        # 6. GO统计Sheet|GO Statistics Sheet
        if include_summary and parser.parse_goterms:
            ws_go_stats = wb.create_sheet("GO Statistics", sheet_index)
            go_stats = parser.get_go_statistics()
            if go_stats:
                df_go = pd.DataFrame(list(go_stats.items()), columns=['GO Category', 'Count'])
                self._write_dataframe_to_sheet(ws_go_stats, df_go, "GO术语分类统计|GO Term Category Statistics")
            else:
                ws_go_stats.append(["无GO术语数据|No GO term data"])
            sheet_index += 1

        # 保存文件|Save file
        wb.save(output_file)
        self.logger.info(f"Excel报告已生成|Excel report generated: {output_file}")

    def _generate_csv(self, parser, output_prefix: str, include_summary: bool):
        """生成CSV报告|Generate CSV report"""
        self.logger.info("生成CSV报告|Generating CSV report")

        # 1. 详细结果CSV|Detailed Results CSV
        if parser.matches:
            df_matches = parser.get_matches_dataframe()
            matches_file = f"{output_prefix}_detailed_results.csv"
            df_matches.to_csv(matches_file, index=False, encoding='utf-8-sig')
            self.logger.info(f"详细结果CSV已生成|Detailed results CSV generated: {matches_file}")

        # 2. GO注释CSV（展开格式，每个GO一行）|GO Annotation CSV (expanded, one GO per row)
        if parser.parse_goterms:
            df_go_expanded = parser.get_expanded_go_dataframe()
            if df_go_expanded is not None and not df_go_expanded.empty:
                go_file = f"{output_prefix}_go_annotations.csv"
                df_go_expanded.to_csv(go_file, index=False, encoding='utf-8-sig')
                self.logger.info(f"GO注释CSV已生成|GO annotations CSV generated: {go_file}")
            else:
                self.logger.info("无GO注释数据|No GO annotation data")

        # 3. Pathway注释CSV（展开格式，每个Pathway一行）|Pathway Annotation CSV (expanded, one Pathway per row)
        if parser.parse_pathways:
            df_pathway_expanded = parser.get_expanded_pathway_dataframe()
            if df_pathway_expanded is not None and not df_pathway_expanded.empty:
                pathway_file = f"{output_prefix}_pathway_annotations.csv"
                df_pathway_expanded.to_csv(pathway_file, index=False, encoding='utf-8-sig')
                self.logger.info(f"Pathway注释CSV已生成|Pathway annotations CSV generated: {pathway_file}")
            else:
                self.logger.info("无Pathway注释数据|No Pathway annotation data")

        # 4. 汇总CSV|Summary CSV
        if include_summary and parser.summaries:
            df_summary = parser.get_summaries_dataframe()
            df_summary = df_summary.sort_values('Total Matches', ascending=False)
            summary_file = f"{output_prefix}_protein_summary.csv"
            df_summary.to_csv(summary_file, index=False, encoding='utf-8-sig')
            self.logger.info(f"汇总CSV已生成|Summary CSV generated: {summary_file}")

        # 5. 数据库统计CSV|Database Statistics CSV
        if include_summary:
            db_stats = parser.get_database_statistics()
            df_db = pd.DataFrame(list(db_stats.items()), columns=['Database', 'Match Count'])
            db_file = f"{output_prefix}_database_stats.csv"
            df_db.to_csv(db_file, index=False, encoding='utf-8-sig')
            self.logger.info(f"数据库统计CSV已生成|Database stats CSV generated: {db_file}")

    def _write_dataframe_to_sheet(self, worksheet, dataframe, title: str):
        """将DataFrame写入工作表并格式化|Write DataFrame to worksheet with formatting"""
        # 写入标题|Write title
        worksheet['A1'] = title
        worksheet['A1'].font = Font(bold=True, size=14)

        # 合并标题单元格|Merge title cells
        end_col = len(dataframe.columns)
        end_col_letter = chr(64 + end_col) if end_col <= 26 else chr(64 + (end_col - 1) // 26) + chr(64 + (end_col - 1) % 26 + 1)
        worksheet.merge_cells(f'A1:{end_col_letter}1')

        # 写入数据|Write data
        for r_idx, row in enumerate(dataframe_to_rows(dataframe, index=False, header=True), 2):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)

                # 表头格式|Header format
                if r_idx == 2:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color='CCE5FF', end_color='CCE5FF', fill_type='solid')
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # 数据格式|Data format
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)

        # 自动调整列宽|Auto-adjust column width
        for col_idx in range(1, len(dataframe.columns) + 1):
            max_length = 0

            for row in worksheet.iter_cols(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    try:
                        if cell.value:
                            cell_length = len(str(cell.value))
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[chr(64 + col_idx)].width = adjusted_width

        # 冻结首行|Freeze first row (after title)
        worksheet.freeze_panes = 'B3'
