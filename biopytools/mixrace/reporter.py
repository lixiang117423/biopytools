"""mixrace 报告层(v0.4 三分支)|mixrace reporting (three-branch verdict).

判读汇总表(TSV/HTML)、单样本 md 报告(证据链+实验建议)、自包含 HTML(关键图内嵌)。
HTML 展示层规范:完整文档骨架(charset 防乱码)、页头判读构成条+总览卡片、
verdict 徽章(色点+文字双编码,不单独依赖颜色)、汇总表横滚容器+粘性表头/首列+
data-v 数值排序、全部图中文标题+图注、证据链搜索/展开、打印样式、零外链零依赖。
|Summary tables, per-sample evidence-chain reports, self-contained HTML.
Presentation: full-document skeleton, header composition strip + stat cards,
badges (color+label, never color alone), sortable sticky table with raw-value
data-v attributes, CN-titled figure gallery, evidence search/expand, print CSS.
"""
import base64
import datetime
import html as _html
from typing import Tuple

_SUMMARY_COLS = ["sample", "verdict", "advice", "het_rate", "robust_rate",
                 "shared_only_rate", "top_partner", "mix_proportion",
                 "het_rate_after_hotspot", "dp_ratio", "host_rate",
                 "pathogen_map_rate", "contamination_rate", "mean_depth", "breadth_1x"]

_VERDICT_CN = {"pure": "纯菌", "divergent": "优势菌株/参考差异型",
               "contaminated": "混杂菌株", "uncertain": "不确定"}

_SUBTAG_CN = {"mild": "轻度"}   # subtag 枚举值→显示名|subtag enum -> display label

# verdict 展示规格:CVD 校验后定稿——精确色只用于色点/构成条段(与 matplotlib 图
# 一致),文字用同色相加深档保 WCAG 对比,浅底做徽章背景|display spec per verdict:
# exact hue for dots/strip segments (matches figures), darkened ink for text,
# light tint for badge background.
_VERDICT_STYLE = {
    "pure":         {"color": "#2e7d32", "ink": "#1b5e20", "bg": "#e9f3ea", "cn": "纯菌"},
    "divergent":    {"color": "#ef6c00", "ink": "#a04a00", "bg": "#fdeede",
                     "cn": "优势菌株/参考差异型"},
    "contaminated": {"color": "#c62828", "ink": "#b71c1c", "bg": "#fbeaea", "cn": "混杂菌株"},
    "uncertain":    {"color": "#9e9e9e", "ink": "#616161", "bg": "#f0f1f1", "cn": "不确定"},
}
_VERDICT_ORDER = ["pure", "divergent", "contaminated", "uncertain"]

# 数值列(右对齐+data-v 原始值供 JS 排序)|numeric cols (right-aligned + raw data-v)
_NUM_COLS = {"het_rate", "robust_rate", "shared_only_rate", "mix_proportion",
             "het_rate_after_hotspot", "dp_ratio", "host_rate",
             "pathogen_map_rate", "contamination_rate", "mean_depth", "breadth_1x"}

_METRIC_EXPLAIN = {
    "het_rate": "总杂合率:单倍体每个位点本该纯合,出现杂合=混合或错误。这是判读的主指标,像混合比例的指纹。",
    "robust_rate": "稳健杂合率:只统计 altAD>=5 且 altfrac>=0.2 的杂合位点(排除低深度测序错误)后的杂合率。",
    "shared_only_rate": "shared-only 杂合率:只统计 ALT 在其他样品也出现的杂合位点(排除样品特异噪声)。",
    "top_partner": "混合伴侣:在我的杂合位点上携带 ALT 最多的样品——相当于在群体里找与你共享变异的『另一半』。",
    "mix_proportion": "成分推断:若判混杂,样品约等于 该比例x伴侣型 + 余量x参考型 的混合。",
    "het_rate_after_hotspot": "排除热点后杂合率:剔除所有混杂样品共享的高差异窗口后剩下的杂合率,仍高则混杂证据坚实。",
    "dp_ratio": "DP检验:杂合位点深度/纯合位点深度。>1.5 说明杂合位点深度反而更高,是真实混合信号而非低覆盖错误。",
    "host_rate": "寄主reads占比:比对上寄主基因组的reads比例,代表样品里植物DNA污染程度。",
    "pathogen_map_rate": "病原mapping率:剔除寄主后比对上病原基因组的reads比例,代表病原数据质量。",
    "contamination_rate": "污染reads占比:既没比对上寄主、也没比对上病原基因组的reads比例,可能来自其他微生物或低质量reads。",
    "mean_depth": "平均测序深度:每个位置平均被测了多少遍。",
    "breadth_1x": "覆盖广度:基因组至少被测到1次的碱基占比;广度低说明有大片区域没测到。",
}

_EVIDENCE_KEYS = ["het_rate", "robust_rate", "shared_only_rate", "top_partner",
                  "mix_proportion", "het_rate_after_hotspot", "dp_ratio",
                  "host_rate", "pathogen_map_rate", "contamination_rate",
                  "mean_depth", "breadth_1x"]

# stem → (中文标题, 一句话图注)|stem -> (CN title, one-line caption)
# 与 figures.py 的图一一对应;未知 stem 回退显示 stem 本身|unknown stems fall
# back to the stem itself as title.
_FIGURE_META = {
    "het_heatmap_100kb": (
        "100kb 窗口杂合率热图",
        "每行一个样品,每列一个 100kb 窗口,颜色越红杂合率越高。纯菌样品应整体接近白色;横向红线带提示该样品大面积高杂合。"),
    "het_heatmap_excl_hotspots": (
        "杂合热图:排除热点前后对比",
        "左=排除前,右=剔除混杂样品共享的高差异窗口后。右侧仍大面积偏红说明高杂合不依赖共享热点,混杂证据更坚实。"),
    "het_genome_overview": (
        "全基因组杂合分布总览",
        "每样品一行,每个点是一个窗口的杂合率。真混合呈贯穿基因组的均匀高带;测序错误或局部噪声只呈孤立散点。"),
    "manhattan_grid": (
        "非纯样本 Manhattan 拼图",
        "只画判读非纯菌的样品,便于聚焦比较各自的窗口杂合分布。"),
    "dist_heatmap": (
        "样本间 SNP 不匹配率热图",
        "格子颜色=两样品间 SNP 不匹配率,越红差异越大。混杂样品与其混合伴侣通常在此矩阵中最接近。"),
    "pca_3view": (
        "PCA 三视图",
        "绿=纯菌,橙=优势菌株/参考差异型,红=混杂。混杂样品常落在纯菌簇与参考/伴侣之间的连线上。"),
    "nj_tree": (
        "NJ 系统发育树",
        "基于样本间 SNP 不匹配率构建。混杂样品倾向于与其混合伴侣聚在同一分支。"),
    "altfrac_dist": (
        "杂合位点 altfrac 分布",
        "横轴=杂合位点上 ALT 碱基所占比例。真实混合呈集中在混合比例附近的峰;测序错误富集在极低 altfrac 一端。"),
    "eval_3panel": (
        "三面板评估",
        "同一批样品在总杂合率、DP>=50 杂合率、稳健杂合率三种口径下的对照。三列同高说明高杂合不是低深度或低质量位点造成的假象。"),
}
# 图版阅读顺序:全景→定位→关系→机理|reading order for the gallery
_FIGURE_ORDER = ["het_heatmap_100kb", "het_heatmap_excl_hotspots",
                 "het_genome_overview", "manhattan_grid", "dist_heatmap",
                 "pca_3view", "nj_tree", "altfrac_dist", "eval_3panel"]


def _html_escape(text: str) -> str:
    return _html.escape(str(text))


def _fmt(key, val) -> str:
    """指标值格式化(None→—;比率→百分比)|format metric value."""
    if val is None or val == "":
        return "—"
    if isinstance(val, str):
        return val
    if key in ("het_rate", "robust_rate", "shared_only_rate", "het_rate_after_hotspot"):
        return f"{val*100:.4f}%" if val < 0.01 else f"{val*100:.2f}%"
    if key in ("host_rate", "pathogen_map_rate", "contamination_rate"):
        return f"{val*100:.2f}%"
    if key == "mix_proportion":
        return f"{val*100:.0f}%"
    if key == "dp_ratio":
        return f"{val:.2f}"
    if key == "mean_depth":
        return f"{val:.1f}x"
    if key == "breadth_1x":
        return f"{val:.2f}%"
    return str(val)


_COLS_CN = {"sample": "样品", "verdict": "判读", "advice": "建议",
            "het_rate": "总杂合率", "robust_rate": "稳健杂合率",
            "shared_only_rate": "shared-only杂合率", "top_partner": "混合伴侣",
            "mix_proportion": "成分推断", "het_rate_after_hotspot": "排除热点后杂合率",
            "dp_ratio": "DP检验", "host_rate": "寄主占比",
            "pathogen_map_rate": "病原mapping率", "contamination_rate": "污染reads",
            "mean_depth": "平均深度", "breadth_1x": "覆盖广度"}


def _format_summary_rows(rows: list) -> list:
    """汇总行格式化(比率→百分比字符串;附 verdict_cn)|format rows (+CN verdict)."""
    disp = []
    for r in rows:
        d = dict(r)
        for k in ("het_rate", "robust_rate", "shared_only_rate",
                  "het_rate_after_hotspot", "host_rate", "pathogen_map_rate",
                  "contamination_rate", "mix_proportion", "dp_ratio",
                  "mean_depth", "breadth_1x"):
            if k in d:
                d[k] = _fmt(k, d[k])
        # TSV 保留机器可读原始键;中文显示名仅给 HTML|TSV keeps raw keys; CN only for HTML
        d["verdict_cn"] = f"{_VERDICT_CN.get(r.get('verdict'), r.get('verdict', ''))}" \
                          f"{_SUBTAG_CN.get(r.get('subtag', ''), r.get('subtag', ''))}"
        disp.append(d)
    return disp


def _badge(verdict: str, subtag: str = "") -> str:
    """verdict 徽章:色点+文字双编码(CVD 安全,不单独依赖颜色)|color+label badge."""
    v = _VERDICT_STYLE.get(str(verdict), _VERDICT_STYLE["uncertain"])
    sub = _SUBTAG_CN.get(str(subtag), str(subtag)) if subtag else ""
    label = _html_escape(v["cn"])
    sub_html = f'<span class="sub">{_html_escape(sub)}</span>' if sub else ""
    return (f'<span class="badge v-{_html_escape(str(verdict))}">'
            f'<span class="dot"></span>{label}{sub_html}</span>')


# ---------------------------------------------------------------- 汇总表|table

def _summary_table_fragment(rows: list) -> str:
    """汇总表 HTML 片段(横滚容器+粘性表头/首列+徽章+data-v)|table fragment."""
    disp = _format_summary_rows(rows)
    head = "".join(
        f'<th class="sortable" data-sortable="1" tabindex="0" title="{_html_escape(c)}">'
        f'{_html_escape(_COLS_CN.get(c, c))}<br><span class="en">{_html_escape(c)}</span></th>'
        for c in _SUMMARY_COLS)
    body_rows = []
    for d, r in zip(disp, rows):
        cells = []
        for c in _SUMMARY_COLS:
            if c == "sample":
                cells.append(f'<td class="smp">{_html_escape(str(d.get(c, "")))}</td>')
            elif c == "verdict":
                v = str(r.get("verdict", "uncertain"))
                cells.append(f'<td>{_badge(v, str(r.get("subtag", "") or ""))}</td>')
            elif c == "advice":
                adv = str(d.get(c, ""))
                cells.append(f'<td class="adv" title="{_html_escape(adv)}">'
                             f'{_html_escape(adv)}</td>')
            elif c == "top_partner":
                cells.append(f'<td>{_html_escape(str(d.get(c, "")))}</td>')
            else:
                raw = r.get(c)
                dv = f' data-v="{raw}"' if isinstance(raw, (int, float)) else ""
                cells.append(f'<td class="num"{dv}>{_html_escape(str(d.get(c, "")))}</td>')
        body_rows.append(f'<tr data-sample="{_html_escape(str(r.get("sample", "")))}">'
                         + "".join(cells) + "</tr>")
    return ('<div class="tbl-wrap"><table class="sum"><thead><tr>' + head
            + "</tr></thead><tbody>" + "".join(body_rows)
            + "</tbody></table></div>")


def build_summary_table(rows: list) -> Tuple[str, str]:
    """判读汇总表(TSV+独立 HTML 文档)|verdict summary table (tsv + standalone html)."""
    disp = _format_summary_rows(rows)
    tsv = "\t".join(_SUMMARY_COLS) + "\n"
    for d in disp:
        tsv += "\t".join(str(d.get(c, "")) for c in _SUMMARY_COLS) + "\n"
    body = _summary_table_fragment(rows)
    html = _document("判读汇总|verdict summary",
                     f'<header class="page"><span class="eyebrow">MIXRACE</span>'
                     f'<h1>判读汇总|verdict summary</h1>'
                     f'<p class="pagemeta">{len(disp)} 份样品|samples</p></header>'
                     f'<section>{body}</section>')
    return tsv, html


# ------------------------------------------------- 页头总览|header overview

def _verdict_counts(rows: list) -> dict:
    counts = {v: 0 for v in _VERDICT_ORDER}
    for r in rows:
        v = str(r.get("verdict", "uncertain"))
        counts[v] = counts.get(v, 0) + 1
    return counts


def _overview_html(rows: list) -> str:
    """构成条+四张计数卡(签名元素:一眼看清批次构成)|strip + stat cards."""
    counts = _verdict_counts(rows)
    total = max(1, len(rows))
    parts = []
    if rows:
        segs = []
        for v in _VERDICT_ORDER:
            n = counts.get(v, 0)
            if n:
                pct = n / total * 100
                segs.append(f'<div class="compo-seg v-{v}" style="width:{pct:.2f}%" '
                            f'aria-label="{_html_escape(_VERDICT_STYLE[v]["cn"])} {n}"></div>')
        parts.append('<div class="compo-strip" role="img" '
                     f'aria-label="判读构成|verdict composition">' + "".join(segs) + "</div>")
    cards = []
    for v in _VERDICT_ORDER:
        st = _VERDICT_STYLE[v]
        cards.append(f'<div class="stat-card"><span class="dot" style="background:{st["color"]}"></span>'
                     f'<div class="n" data-count="{counts.get(v, 0)}">{counts.get(v, 0)}</div>'
                     f'<div class="lbl">{_html_escape(st["cn"])}</div></div>')
    parts.append('<div class="stat-grid">' + "".join(cards) + "</div>")
    return "".join(parts)


# ------------------------------------------------------------- 图版|figures

def _embed_image(path: str):
    """PNG → base64(单文件HTML内嵌)|PNG to base64 for self-contained HTML."""
    try:
        with open(path, "rb") as fh:
            return base64.b64encode(fh.read()).decode("ascii")
    except OSError:
        return None


def _figure_gallery_html(figures: dict) -> str:
    """全部图按阅读顺序入报告,中文标题+图注|all figures in reading order + captions."""
    stems = [s for s in _FIGURE_ORDER if figures.get(s)]
    stems += [s for s in figures if s not in _FIGURE_ORDER and figures.get(s)]
    parts = []
    for stem in stems:
        b64 = _embed_image(figures.get(stem, ""))
        if not b64:
            continue
        title, caption = _FIGURE_META.get(stem, (stem, ""))
        cap = f"<p>{_html_escape(caption)}</p>" if caption else ""
        parts.append(f'<figure><figcaption><h3>{_html_escape(title)}</h3>{cap}</figcaption>'
                     f'<img src="data:image/png;base64,{b64}" alt="{_html_escape(title)}"></figure>')
    return "".join(parts)


# ----------------------------------------------------------- 证据链|evidence

def _evidence_html(rows: list) -> str:
    """逐样品证据链:搜索+展开按钮+details(中文化指标)|per-sample evidence."""
    parts = ['<div class="toolbar no-print">',
             '<input id="sample-search" type="text" placeholder="搜索样品|filter samples">',
             '<button type="button" class="expand-all">全部展开</button>',
             '<button type="button" class="collapse-all">全部收起</button>',
             "</div>"]
    for r in rows:
        verdict = str(r.get("verdict", "uncertain"))
        sample = _html_escape(str(r.get("sample", "")))
        quick = _html_escape(_fmt("het_rate", r.get("het_rate")) if r.get("het_rate") is not None else "—")
        parts.append(
            f'<details class="ev" data-sample="{sample}">'
            f'<summary><b>{sample}</b> {_badge(verdict, str(r.get("subtag", "") or ""))}'
            f'<span class="quick">总杂合率 {quick}</span></summary>'
            f'<p class="advice">{_html_escape(str(r.get("advice", "")))}</p>'
            f'<p class="rationale">{_html_escape(str(r.get("rationale", "")))}</p>')
        parts.append('<table><thead><tr><th>指标</th><th>值</th><th>解释</th></tr></thead><tbody>')
        for k in _EVIDENCE_KEYS:
            if k in r:
                parts.append(
                    f'<tr><td>{_html_escape(_COLS_CN.get(k, k))}'
                    f'<br><span class="en">{_html_escape(k)}</span></td>'
                    f'<td class="num">{_html_escape(_fmt(k, r.get(k)))}</td>'
                    f'<td class="expl">{_html_escape(_METRIC_EXPLAIN.get(k, ""))}</td></tr>')
        parts.append("</tbody></table></details>")
    return "".join(parts)


# ------------------------------------------------------------ 主报告|report

def build_html_report(title: str, rows: list, figures: dict,
                      verdict_note: str = "") -> str:
    """自包含 HTML(判读表+关键图内嵌+逐样品证据链折叠)|self-contained HTML report.

    verdict_note: 判读口径文案(阈值可配,由调用方按 config 生成防漂移)
    |verdict_note: threshold note built from config by the caller (no drift).
    """
    ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    body = [f'<header class="page">',
            f'<span class="eyebrow">MIXRACE · 混杂评估|CONTAMINATION ASSESSMENT</span>',
            f'<h1>{_html_escape(title)}</h1>',
            f'<p class="pagemeta">{len(rows)} 份样品|samples · 生成时间|generated {ts}</p>',
            _overview_html(rows)]
    if verdict_note:
        body.append(f'<div class="note">{_html_escape(verdict_note)}</div>')
    body.append('</header>')
    body.append('<section id="summary">'
                '<span class="eyebrow">01 · 判读汇总|VERDICT SUMMARY</span>'
                '<h2>判读汇总表</h2>'
                '<p class="sub">点击表头排序;左右滚动查看全部指标|click headers to sort; scroll horizontally</p>'
                + _summary_table_fragment(rows) + "</section>")
    fig_html = _figure_gallery_html(figures)
    if fig_html:
        body.append('<section id="figures">'
                    '<span class="eyebrow">02 · 图版|FIGURES</span>'
                    '<h2>图版</h2>'
                    '<p class="sub">按阅读顺序排列:全景→定位→关系→机理|ordered overview to mechanism</p>'
                    + fig_html + "</section>")
    body.append('<section id="evidence">'
                '<span class="eyebrow">03 · 逐样品证据链|PER-SAMPLE EVIDENCE</span>'
                '<h2>逐样品证据链</h2>'
                '<p class="sub">展开查看每个样品的完整证据与建议|expand for full evidence chain</p>'
                + _evidence_html(rows) + "</section>")
    return _document(title, "".join(body))


# ------------------------------------------------------- 文档骨架|document

def _document(title: str, body_html: str) -> str:
    """完整 HTML 文档(charset/viewport/样式/脚本)|full document skeleton."""
    return ("<!DOCTYPE html>\n<html lang=\"zh-CN\">\n<head>\n"
            "<meta charset=\"utf-8\">\n"
            "<meta name=\"viewport\" content=\"width=device-width, initial-scale=1\">\n"
            f"<title>{_html_escape(title)}</title>\n"
            f"<style>{_CSS}</style>\n</head>\n<body>\n"
            f'<div class="wrap">{body_html}</div>\n'
            f"<script>{_JS}</script>\n</body>\n</html>\n")


def build_sample_report(sample: str, row: dict, figures: dict) -> str:
    """单样本 md 报告(证据链+建议+图)|per-sample markdown (evidence + advice + figures)."""
    verdict = str(row.get("verdict", "uncertain"))
    md = [f"# {sample} 混杂评估报告|{sample} contamination report", ""]
    md.append(f"**判读|Verdict: {_VERDICT_CN.get(verdict, verdict)}"
              f"{_SUBTAG_CN.get(row.get('subtag', ''), row.get('subtag', '') or '')}**")
    md.append("")
    md.append(f"**{row.get('advice', '')}**")
    md.append("")
    md.append("## 依据|Rationale")
    md.append("")
    md.append(str(row.get("rationale", "")))
    md.append("")
    md.append("## 证据链|Evidence")
    md.append("")
    md.append("| 指标<br>Metric | 值<br>Value | 通俗解释<br>Plain words |")
    md.append("|---|---|---|")
    for k in _EVIDENCE_KEYS:
        if k in row:
            md.append(f"| {k} | {_fmt(k, row.get(k))} | {_METRIC_EXPLAIN.get(k, '')} |")
    md.append("")
    if figures:
        md.append("## 图|Figures")
        md.append("")
        for stem, p in figures.items():
            md.append(f"![{stem}]({p})")
            md.append("")
    return "\n".join(md)


_CSS = """
:root{--ink:#1c2b33;--muted:#5c6b73;--line:#dde4e8;--paper:#fff;--panel:#f5f7f8;
--v-pure:#2e7d32;--v-divergent:#ef6c00;--v-contaminated:#c62828;--v-uncertain:#9e9e9e}
*{box-sizing:border-box}
body{margin:0;background:var(--paper);color:var(--ink);
font:15px/1.65 -apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Hiragino Sans GB","Noto Sans CJK SC","Microsoft YaHei",sans-serif}
.wrap{max-width:1180px;margin:0 auto;padding:38px 28px 72px}
header.page h1{font-size:26px;line-height:1.3;margin:0 0 6px;letter-spacing:.01em}
.pagemeta{color:var(--muted);font-size:13px;margin:0}
.eyebrow{display:block;font-size:11px;letter-spacing:.18em;color:var(--muted);margin-bottom:4px}
section{margin-top:46px}
h2{font-size:19px;margin:0 0 4px}
h3{font-size:15px;margin:0}
.sub{color:var(--muted);font-size:13px;margin:0 0 14px}
.en{color:var(--muted);font-weight:400;font-size:11px}
.compo-strip{display:flex;height:10px;border-radius:5px;overflow:hidden;gap:2px;margin:18px 0 14px}
.compo-seg{min-width:2px}
.compo-seg.v-pure{background:var(--v-pure)}
.compo-seg.v-divergent{background:var(--v-divergent)}
.compo-seg.v-contaminated{background:var(--v-contaminated)}
.compo-seg.v-uncertain{background:var(--v-uncertain)}
.stat-grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:12px}
.stat-card{border:1px solid var(--line);border-radius:8px;padding:14px 16px;background:var(--panel)}
.stat-card .dot{display:inline-block;width:9px;height:9px;border-radius:50%;margin-bottom:6px}
.stat-card .n{font-size:26px;font-weight:700;line-height:1.1;font-variant-numeric:tabular-nums}
.stat-card .lbl{color:var(--muted);font-size:12.5px;margin-top:2px}
.note{border-left:3px solid var(--v-uncertain);background:var(--panel);padding:10px 14px;
font-size:13.5px;color:var(--muted);border-radius:0 6px 6px 0;margin-top:24px}
.tbl-wrap{overflow:auto;border:1px solid var(--line);border-radius:8px;max-height:72vh}
table{border-collapse:separate;border-spacing:0;width:100%;margin:0;font-size:13px;
font-variant-numeric:tabular-nums}
thead th{position:sticky;top:0;z-index:2;background:var(--panel);text-align:left;
font-weight:600;border-bottom:1px solid var(--line);padding:8px 10px;white-space:nowrap}
thead th.sortable{cursor:pointer;user-select:none}
thead th.sortable:hover{background:#e9edf0}
thead th.sortable:focus-visible{outline:2px solid var(--v-uncertain);outline-offset:-2px}
thead th:first-child{z-index:3}
tbody td{padding:7px 10px;border-bottom:1px solid var(--line);white-space:nowrap;background:var(--paper)}
tbody tr:nth-child(even) td{background:#fafbfc}
tbody td:first-child{position:sticky;left:0;z-index:1;background:var(--paper)}
tbody tr:nth-child(even) td:first-child{background:#fafbfc}
table tbody tr:hover td{background:#eef2f4}
td.num{text-align:right;font-family:ui-monospace,SFMono-Regular,Consolas,"Liberation Mono",monospace}
td.adv{max-width:230px;overflow:hidden;text-overflow:ellipsis}
.badge{display:inline-flex;align-items:center;gap:6px;padding:2px 9px 2px 7px;
border-radius:999px;font-weight:600;font-size:12.5px;white-space:nowrap}
.badge .dot{width:8px;height:8px;border-radius:50%;flex:none}
.badge .sub{font-weight:400;font-size:11px;opacity:.85}
.badge.v-pure{background:#e9f3ea;color:#1b5e20}.badge.v-pure .dot{background:var(--v-pure)}
.badge.v-divergent{background:#fdeede;color:#a04a00}.badge.v-divergent .dot{background:var(--v-divergent)}
.badge.v-contaminated{background:#fbeaea;color:#b71c1c}.badge.v-contaminated .dot{background:var(--v-contaminated)}
.badge.v-uncertain{background:#f0f1f1;color:#616161}.badge.v-uncertain .dot{background:var(--v-uncertain)}
figure{margin:26px 0 0;border:1px solid var(--line);border-radius:8px;padding:14px 16px 10px}
figure img{max-width:100%;height:auto;display:block;margin:10px auto 4px}
figcaption p{margin:4px 0 0;color:var(--muted);font-size:13px}
.toolbar{display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px}
.toolbar input{padding:7px 12px;border:1px solid var(--line);border-radius:6px;font-size:13.5px;width:240px}
.toolbar button{padding:7px 12px;border:1px solid var(--line);background:var(--panel);
border-radius:6px;font-size:13px;cursor:pointer;color:var(--ink)}
.toolbar button:hover{background:#e9edf0}
details.ev{border:1px solid var(--line);border-radius:8px;margin:8px 0;padding:0 14px;background:var(--paper)}
details.ev[open]{background:var(--panel)}
details.ev summary{cursor:pointer;padding:10px 0;font-weight:600;list-style:none;
display:flex;align-items:center;gap:10px;flex-wrap:wrap}
details.ev summary::-webkit-details-marker{display:none}
details.ev summary .quick{color:var(--muted);font-weight:400;font-size:12.5px;margin-left:auto}
details.ev .advice{margin:2px 0 6px;font-size:13.5px}
details.ev .rationale{margin:0 0 8px;color:var(--muted);font-size:13px}
details.ev table{margin:4px 0 14px;font-size:12.5px}
details.ev td,details.ev th{padding:6px 10px;border-bottom:1px solid var(--line);
text-align:left;vertical-align:top;white-space:normal}
details.ev td.num{white-space:nowrap}
details.ev td.expl{color:var(--muted)}
@media print{
.toolbar{display:none}
.wrap{max-width:none;padding:0}
.tbl-wrap{max-height:none;overflow:visible;border:none}
figure{break-inside:avoid}
details.ev{break-inside:avoid}
}
"""

_JS = """
function sortTable(th){
  var table=th.closest('table'),tbody=table.querySelector('tbody');
  var idx=Array.prototype.indexOf.call(th.parentNode.children,th);
  var asc=th.getAttribute('data-asc')!=='1';
  Array.prototype.forEach.call(th.parentNode.children,function(o){
    o.removeAttribute('data-asc');o.removeAttribute('aria-sort');});
  th.setAttribute('data-asc',asc?'1':'0');
  th.setAttribute('aria-sort',asc?'ascending':'descending');
  var rows=Array.prototype.slice.call(tbody.querySelectorAll('tr'));
  rows.sort(function(a,b){
    var ca=a.children[idx],cb=b.children[idx];
    var va=ca.getAttribute('data-v'),vb=cb.getAttribute('data-v');
    if(va!==null&&vb!==null&&va!==''&&vb!==''){return (parseFloat(va)-parseFloat(vb))*(asc?1:-1);}
    return ca.textContent.localeCompare(cb.textContent,'zh')*(asc?1:-1);});
  rows.forEach(function(r){tbody.appendChild(r);});
}
(function(){
  // 表头排序接线(点击+键盘)|wire header sort (click + keyboard)
  document.querySelectorAll('th.sortable').forEach(function(th){
    th.addEventListener('click',function(){sortTable(th);});
    th.addEventListener('keydown',function(e){
      if(e.key==='Enter'||e.key===' '){e.preventDefault();sortTable(th);}});});
  var box=document.getElementById('sample-search');
  if(box){box.addEventListener('input',function(){
    var q=box.value.trim().toLowerCase();
    document.querySelectorAll('details.ev').forEach(function(d){
      var hit=!q||d.getAttribute('data-sample').toLowerCase().indexOf(q)>=0;
      d.style.display=hit?'':'none';if(q&&hit){d.open=true;}});
    document.querySelectorAll('table.sum tbody tr').forEach(function(tr){
      var hit=!q||tr.getAttribute('data-sample').toLowerCase().indexOf(q)>=0;
      tr.style.display=hit?'':'none';});});}
  var openAll=function(){document.querySelectorAll('details.ev').forEach(function(d){d.open=true;});};
  var oa=document.querySelector('button.expand-all');if(oa){oa.addEventListener('click',openAll);}
  var ca=document.querySelector('button.collapse-all');
  if(ca){ca.addEventListener('click',function(){
    document.querySelectorAll('details.ev').forEach(function(d){d.open=false;});});}
  window.addEventListener('beforeprint',openAll);
})();
"""


def write_summary_excel(rows: list, path: str) -> None:
    """判读汇总 Excel:中文表头 sheet + 英文表头 sheet|CN-header + EN-header sheets.

    中文 sheet(判读汇总)表头/判读列用中文显示名;英文 sheet(summary)保留原始
    机器可读键;两 sheet 数据与格式化一致,表头加粗+冻结首行+自适应列宽。
    |CN sheet uses display names; EN sheet keeps raw keys; bold frozen headers.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    disp = _format_summary_rows(rows)
    wb = Workbook()

    def _fill(ws, headers, verdict_of_row):
        ws.append(headers)
        for d in disp:
            ws.append([verdict_of_row(d) if c == "verdict" else str(d.get(c, ""))
                       for c in _SUMMARY_COLS])
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for j, h in enumerate(headers, start=1):
            ws.column_dimensions[ws.cell(1, j).column_letter].width = max(10, len(str(h)) * 2.2)

    ws_cn = wb.active
    ws_cn.title = "判读汇总"
    _fill(ws_cn, [_COLS_CN.get(c, c) for c in _SUMMARY_COLS],
          lambda d: str(d.get("verdict_cn", d.get("verdict", ""))))
    ws_en = wb.create_sheet("summary")
    _fill(ws_en, list(_SUMMARY_COLS), lambda d: str(d.get("verdict", "")))
    wb.save(path)
